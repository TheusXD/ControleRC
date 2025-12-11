import streamlit as st
import pandas as pd
import os
import time
import hashlib
from datetime import datetime, timedelta
import io
import firebase_admin
from firebase_admin import credentials, firestore, exceptions
import plotly.express as px
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, ValidationError, EmailStr
from typing import List, Dict, Any, Optional, Tuple
import re
import logging
import json
import base64
import uuid
import bleach
import requests

@st.cache_resource
def get_db_service(creds):
    """Cria a conexão com o banco apenas uma vez."""
    return FirebaseService(creds)

@st.cache_data(ttl=300)  # Cache de 5 minutos para tabelas gerais
def get_cached_docs(_db_service, collection_name):
    """Lê dados do banco e guarda na memória."""
    # O underline (_) no nome do argumento evita erro de hash do Streamlit
    return _db_service.get_docs(collection_name)

@st.cache_data(ttl=600)  # Cache de 10 minutos para lista de usuários
def get_cached_users_list(_db_service):
    """Carrega a lista de usuários apenas uma vez a cada 10 min."""
    df = get_cached_docs(_db_service, "users")
    if not df.empty and 'username' in df.columns:
        return ["Ninguém"] + sorted(df['username'].unique().tolist())
    return ["Ninguém"]

def clear_cache():
    """Limpa a memória para mostrar dados novos após salvar."""
    st.cache_data.clear()

# Configurar o logging para monitorizar a aplicação
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Constantes para o sistema de comentários
MAX_COMMENTS_IN_DOCUMENT = 20
MAX_COMMENTS_DISPLAY = 50


# --- SOLUÇÃO DA TRANSAÇÃO: FUNÇÃO STANDALONE ---
@firestore.transactional
def _atomic_add_and_update_standalone(transaction, db, add_col, add_data, update_col, update_id, update_data):
    """
    Função transacional standalone que não depende de 'self' para evitar conflitos com o decorador.
    """
    update_doc_ref = db.collection(update_col).document(update_id)
    snapshot = update_doc_ref.get(transaction=transaction)

    if not snapshot.exists:
        raise ValueError(f"Documento {update_id} não encontrado em {update_col}")

    new_doc_ref = db.collection(add_col).document()
    transaction.set(new_doc_ref, add_data)
    transaction.update(update_doc_ref, update_data)


# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Controle de Compras", layout="wide")


# -----------------------------------------------------------------------------
# 1. MODELS (VALIDAÇÃO DE DADOS COM PYDANTIC)
# -----------------------------------------------------------------------------

class Demanda(BaseModel):
    solicitante_demanda: str = Field(..., min_length=1)
    descricao_necessidade: str = Field(..., min_length=5)
    tipo: str = Field(...)
    categoria: str = Field(..., min_length=1)
    anexo: Optional[Dict[str, Any]] = Field(default=None)
    status_demanda: str = Field(default="Aberta")
    created_at: datetime = Field(default_factory=datetime.now)
    updated_at: datetime = Field(default_factory=datetime.now)
    assigned_to: Optional[str] = Field(default=None)
    closed_at: Optional[datetime] = Field(default=None)
    historico: List[str] = Field(default_factory=list)
    comentarios: List[Dict[str, Any]] = Field(default_factory=list)
    using_comment_subcollection: bool = Field(default=False)


class Requisicao(BaseModel):
    solicitante: str = Field(..., min_length=1)
    demanda_id: Optional[str] = Field(default=None)
    numero_rc: Optional[str] = Field(default=None)
    valor: float = Field(..., gt=0)
    status: str = Field(default="Aberto")
    created_at: datetime = Field(default_factory=datetime.now)
    updated_at: datetime = Field(default_factory=datetime.now)
    historico: List[str] = Field(default_factory=list)
    comentarios: List[Dict[str, Any]] = Field(default_factory=list)
    using_comment_subcollection: bool = Field(default=False)


class Pedido(BaseModel):
    requisicao_id: str = Field(..., min_length=1)
    solicitante: str = Field(..., min_length=1)
    valor: float = Field(..., gt=0)
    numero_pedido: Optional[str] = Field(default=None)
    status: str = Field(default="Em Processamento")
    created_at: datetime = Field(default_factory=datetime.now)
    updated_at: datetime = Field(default_factory=datetime.now)
    observacao: Optional[str] = Field(default=None)
    data_entrega: Optional[datetime] = Field(default=None)
    email_notificacao: Optional[str] = Field(default=None)
    anexo_email: Optional[Dict[str, str]] = Field(default=None)
    historico: List[str] = Field(default_factory=list)
    comentarios: List[Dict[str, Any]] = Field(default_factory=list)
    using_comment_subcollection: bool = Field(default=False)


class User(BaseModel):
    username: str = Field(..., min_length=1)
    email: EmailStr = Field(...)
    role: str = Field(...)
    status: str = Field(...)
    department: Optional[str] = Field(default=None)
    created_at: datetime = Field(default_factory=datetime.now)
    permissions: List[str] = Field(default_factory=list)


class SolicitacaoCadastro(BaseModel):
    solicitante: str = Field(...)
    descricao: str = Field(..., min_length=5)
    status: str = Field(default="Cadastrando")
    codigo_item_final: Optional[str] = Field(default=None)
    created_at: datetime = Field(default_factory=datetime.now)
    updated_at: datetime = Field(default_factory=datetime.now)
    comentarios: List[Dict[str, Any]] = Field(default_factory=list)


# -----------------------------------------------------------------------------
# 2. SERVICES (LÓGICA DE NEGÓCIOS E ACESSO A DADOS)
# -----------------------------------------------------------------------------

class FirebaseService:
    def add_attachment(self, collection: str, doc_id: str, file_data: dict) -> bool:
        """Adiciona um arquivo à lista de anexos do documento."""
        try:
            doc_ref = self.db.collection(collection).document(doc_id)
            # Usa arrayUnion para adicionar sem apagar os existentes
            doc_ref.update({
                "attachments": firestore.ArrayUnion([file_data])
            })
            clear_cache()
            return True
        except Exception as e:
            logger.error(f"Erro ao anexar arquivo: {e}")
            return False

    def remove_attachment(self, collection: str, doc_id: str, file_data: dict) -> bool:
        """Remove um arquivo da lista."""
        try:
            doc_ref = self.db.collection(collection).document(doc_id)
            doc_ref.update({
                "attachments": firestore.ArrayRemove([file_data])
            })
            clear_cache()
            return True
        except Exception as e:
            logger.error(f"Erro ao remover anexo: {e}")
            return False

    def __init__(self, creds: Dict[str, Any]):
        if not firebase_admin._apps:
            cred_dict = creds
            cred_dict['private_key'] = cred_dict['private_key'].replace('\\n', '\n')
            cert = credentials.Certificate(cred_dict)
            firebase_admin.initialize_app(cert)
        self.db = firestore.client()
        logger.info("Firebase Service inicializado.")

    def log_action(self, action: str, username: str, details: Optional[Dict] = None):
        try:
            log_data = {"timestamp": datetime.now(), "action": action, "username": username, "details": details or {}}
            self.db.collection("audit_logs").add(log_data)
        except Exception as e:
            logger.error(f"Falha ao registrar ação no log de auditoria: {e}", exc_info=True)

    def get_doc(self, collection: str, doc_id: str) -> Optional[Dict[str, Any]]:
        try:
            doc = self.db.collection(collection).document(doc_id).get()
            if doc.exists:
                doc_data = doc.to_dict();
                doc_data['id'] = doc.id
                return doc_data
            return None
        except Exception as e:
            logger.error(f"Erro ao buscar documento {doc_id} de '{collection}': {e}", exc_info=True);
            return None

    def get_docs(self, collection: str, filters: Optional[List[Tuple]] = None) -> pd.DataFrame:
        try:
            query = self.db.collection(collection)
            if filters:
                for f in filters: query = query.where(filter=firestore.FieldFilter(f[0].strip(), f[1], f[2]))
            if collection in ["audit_logs", "notifications"]:
                query = query.order_by("timestamp", direction=firestore.Query.DESCENDING)
            docs = query.stream()
            data = [doc.to_dict() | {'id': doc.id} for doc in docs]
            df = pd.DataFrame(data) if data else pd.DataFrame()
            if 'created_at' in df.columns and collection not in ["audit_logs", "notifications"]:
                df['created_at'] = pd.to_datetime(df['created_at'])
                df = df.sort_values(by='created_at', ascending=False)
            return df
        except Exception as e:
            logger.error(f"Erro ao obter dados de '{collection}': {e}", exc_info=True)
            st.error(f"Erro ao buscar dados de '{collection}': {e}");
            return pd.DataFrame()

    def add_doc(self, collection: str, data: Dict[str, Any]) -> Optional[str]:
        """ Adiciona um documento e retorna seu ID em caso de sucesso. """
        try:
            _, new_doc_ref = self.db.collection(collection).add(data)
            clear_cache()  # <--- ADICIONADO: Limpa cache após criar
            return new_doc_ref.id
        except Exception as e:
            logger.error(f"Erro ao adicionar documento a '{collection}': {e}", exc_info=True)
            st.error(f"Erro ao adicionar em '{collection}': {e}")
            return None

    def _format_value_for_history(self, value) -> str:
        if value is None: return "Vazio"
        if isinstance(value, datetime): return value.strftime('%d/%m/%Y %H:%M')
        if isinstance(value, (dict, list)): return f"[{type(value).__name__} com {len(value)} item(s)]"
        if isinstance(value, bool): return "Sim" if value else "Não"
        if isinstance(value, (int, float)): return str(value)
        if isinstance(value, str):
            if len(value) > 100: return value[:97] + "..."
            return value
        return str(value)[:100]

    def update_doc(self, collection: str, doc_id: str, new_data: Dict[str, Any], username: str) -> bool:
        try:
            doc_ref = self.db.collection(collection).document(doc_id)
            current_doc = doc_ref.get()
            if not current_doc.exists:
                st.error("Documento não encontrado para atualização.");
                return False
            old_data = current_doc.to_dict()
            history_log = old_data.get('historico', [])
            now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
            for key, value in new_data.items():
                if key in ['comentarios', 'historico', 'updated_at', 'using_comment_subcollection']: continue
                old_value = old_data.get(key)
                if old_value != value:
                    old_formatted = self._format_value_for_history(old_value)
                    new_formatted = self._format_value_for_history(value)
                    field_name = key.replace('_', ' ').capitalize()
                    history_log.append(
                        f"'{field_name}' alterado de '{old_formatted}' para '{new_formatted}' por {username} em {now_str}")
            if 'historico' in old_data: new_data['historico'] = history_log
            new_data['updated_at'] = datetime.now()
            doc_ref.update(new_data)
            clear_cache()
            return True
        except Exception as e:
            logger.error(f"Erro ao atualizar documento ID: {doc_id} em '{collection}': {e}", exc_info=True)
            st.error(f"Erro ao atualizar em '{collection}': {e}");
            return False

    def delete_doc(self, collection: str, doc_id: str) -> bool:
        try:
            self.db.collection(collection).document(doc_id).delete();
            clear_cache()
            return True
        except Exception as e:
            logger.error(f"Erro ao excluir documento ID: {doc_id} de '{collection}': {e}", exc_info=True)
            st.error(f"Erro ao excluir de '{collection}': {e}");
            return False

    def add_and_update_atomically(self, add_col, add_data, update_col, update_id, update_data) -> bool:
        try:
            transaction = self.db.transaction()
            _atomic_add_and_update_standalone(
                transaction, self.db, add_col, add_data, update_col, update_id, update_data
            )
            clear_cache()
            return True
        except ValueError as e:
            logger.error(f"Falha na transação atômica (ValueError): {e}", exc_info=True)
            st.error(f"Não foi possível concluir a operação: {e}")
            return False
        except Exception as e:
            logger.error(f"Falha na transação atômica (Exception): {e}", exc_info=True)
            st.error(f"Ocorreu um erro de consistência de dados: {e}")
            return False

    def restore_from_backup_data(self, backup_data: dict) -> bool:
        def convert_strings_to_types(obj):
            if isinstance(obj, dict): return {key: convert_strings_to_types(value) for key, value in obj.items()}
            if isinstance(obj, list): return [convert_strings_to_types(element) for element in obj]
            if isinstance(obj, str):
                try:
                    return datetime.fromisoformat(obj)
                except (ValueError, TypeError):
                    try:
                        if obj.isdigit(): return int(obj)
                    except ValueError:
                        pass
                    try:
                        return float(obj)
                    except ValueError:
                        return obj
            return obj

        try:
            collections = ["users", "demandas", "requisicoes", "pedidos", "audit_logs", "notifications"]
            for collection in collections:
                for doc in self.db.collection(collection).stream(): doc.reference.delete()
            for collection, records in backup_data.items():
                if collection in collections:
                    for record in records:
                        if 'id' in record: record.pop('id')
                        processed_record = convert_strings_to_types(record)
                        if collection == 'users':
                            if 'password' in processed_record and isinstance(processed_record.get('password'), str):
                                processed_record['password'] = base64.b64decode(processed_record['password'])
                            if 'salt' in processed_record and isinstance(processed_record.get('salt'), str):
                                processed_record['salt'] = base64.b64decode(processed_record['salt'])
                        self.db.collection(collection).add(processed_record)
            logger.info("Dados restaurados com sucesso");
            return True
        except Exception as e:
            logger.error(f"Falha ao restaurar dados: {e}", exc_info=True);
            return False

    def add_comment_to_subcollection(self, collection: str, doc_id: str, comment_data: Dict) -> bool:
        try:
            self.db.collection(collection).document(doc_id).collection('comments').add(comment_data);
            return True
        except Exception as e:
            logger.error(f"Erro ao adicionar comentário à subcoleção: {e}");
            return False

    def get_comments_from_subcollection(self, collection: str, doc_id: str, limit: int = 50) -> List[Dict]:
        try:
            comments_ref = self.db.collection(collection).document(doc_id).collection('comments')
            comments = comments_ref.order_by('timestamp', direction=firestore.Query.DESCENDING).limit(limit).stream()
            return [{'id': doc.id, **doc.to_dict()} for doc in comments]
        except Exception as e:
            logger.error(f"Erro ao obter comentários da subcoleção: {e}");
            return []

    def migrate_comments_to_subcollection(self, collection: str, doc_id: str) -> bool:
        try:
            doc_ref = self.db.collection(collection).document(doc_id)
            doc = doc_ref.get()
            if not doc.exists: return False
            doc_data = doc.to_dict()
            comments = doc_data.get('comentarios', [])
            if not comments: return True
            for comment in comments:
                if not comment.get('id'): comment['id'] = str(uuid.uuid4())
                self.add_comment_to_subcollection(collection, doc_id, comment)
            doc_ref.update({'comentarios': [], 'using_comment_subcollection': True})
            logger.info(f"Migrados {len(comments)} comentários para subcoleção em {collection}/{doc_id}");
            return True
        except Exception as e:
            logger.error(f"Erro ao migrar comentários: {e}");
            return False


class AuthService:
    SESSION_TIMEOUT_MINUTES = 30
    PERMISSION_REFRESH_SECONDS = 300
    LOGIN_ATTEMPTS_LIMIT = 5
    LOCKOUT_MINUTES = 5

    def __init__(self, db_service: FirebaseService):
        self.db = db_service

    def _hash_password(self, password: str, salt: Optional[bytes] = None) -> Tuple[bytes, bytes]:
        if salt is None: salt = os.urandom(16)
        return hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000), salt

    def _check_password(self, stored_password, salt, provided_password: str) -> bool:
        if salt is None or stored_password is None: return False
        if isinstance(salt, str):
            try:
                salt = base64.b64decode(salt)
            except:
                return False
        if isinstance(stored_password, str):
            try:
                stored_password = base64.b64decode(stored_password)
            except:
                return False
        return stored_password == hashlib.pbkdf2_hmac('sha256', provided_password.encode('utf-8'), salt, 100000)

    def _validate_password_strength(self, password: str) -> bool:
        return len(password) >= 8 and re.search(r"[A-Z]", password) and re.search(r"[a-z]", password) and re.search(
            r"[0-9]", password)

    def register_user(self, username, email, password, is_gestor):
        if not self._validate_password_strength(password):
            st.error("Senha deve ter 8+ caracteres, com maiúscula, minúscula e número.");
            return
        if not self.db.get_docs("users", [("username", "==", username)]).empty:
            st.error("Este nome de usuário já existe.");
            return
        if not self.db.get_docs("users", [("email", "==", email)]).empty:
            st.error("Este e-mail já está em uso.");
            return
        role = "admin" if self.db.get_docs("users").empty else "gestor" if is_gestor else "user"
        status = "active" if role == "admin" or not is_gestor else "pending"
        hashed_pw, salt = self._hash_password(password)
        try:
            user_data = User(username=username, email=email, role=role, status=status).model_dump()
            user_data.update({"password": hashed_pw, "salt": salt})
            if self.db.add_doc("users", user_data):
                self.db.log_action("User Registered", username, {"email": email, "role": role})
                st.success(f"Usuário '{username}' registrado como '{role}'. Status: {status}")
                time.sleep(2);
                st.session_state.page = "Login";
                st.rerun()
        except ValidationError as e:
            st.error(f"Erro de validação: {e}")

    def login_user(self, username, password):
        if 'login_attempts' not in st.session_state: st.session_state.login_attempts = {}
        if username not in st.session_state.login_attempts:
            st.session_state.login_attempts[username] = {'count': 0, 'lockout_until': None}
        user_attempts = st.session_state.login_attempts[username]
        if user_attempts['lockout_until'] and datetime.now() < user_attempts['lockout_until']:
            remaining_time = (user_attempts['lockout_until'] - datetime.now()).seconds
            st.error(f"Muitas tentativas falhas. Tente novamente em {remaining_time} segundos.");
            return
        user_df = self.db.get_docs("users", [("username", "==", username)])
        if not user_df.empty:
            user_data_row = user_df.iloc[0]
            if user_data_row['status'] == 'pending':
                st.warning("Sua conta está aguardando aprovação.")
            elif self._check_password(user_data_row['password'], user_data_row['salt'], password):
                st.session_state.logged_in = True;
                st.session_state.user_data = user_data_row.to_dict()
                st.session_state.username = user_data_row['username'];
                st.session_state.role = user_data_row['role']
                st.session_state.last_activity = time.time();
                st.session_state.last_permission_refresh = time.time()
                user_attempts['count'] = 0;
                user_attempts['lockout_until'] = None
                self.db.log_action("User Login", username);
                st.rerun()
            else:
                user_attempts['count'] += 1
                if user_attempts['count'] >= self.LOGIN_ATTEMPTS_LIMIT:
                    user_attempts['lockout_until'] = datetime.now() + timedelta(minutes=self.LOCKOUT_MINUTES)
                    st.error(f"Usuário ou senha incorretos. Conta bloqueada por {self.LOCKOUT_MINUTES} minutos.")
                    self.db.log_action("Login Failed - Account Locked", username)
                else:
                    st.error("Usuário ou senha incorretos.")
        else:
            st.error("Usuário ou senha incorretos.")

    def check_session_timeout(self):
        if 'last_activity' in st.session_state:
            if time.time() - st.session_state.last_activity > self.SESSION_TIMEOUT_MINUTES * 60:
                self.db.log_action("Session Timeout", st.session_state.get('username', 'unknown'))
                for key in list(st.session_state.keys()): del st.session_state[key]
                st.warning("Sessão expirada por inatividade. Faça login novamente.");
                time.sleep(3);
                st.rerun()
            if time.time() - st.session_state.get('last_permission_refresh', 0) > self.PERMISSION_REFRESH_SECONDS:
                user_id = st.session_state.user_data.get('id')
                if user_id:
                    fresh_user_data = self.db.get_doc("users", user_id)
                    if fresh_user_data:
                        st.session_state.user_data = fresh_user_data
                        st.session_state.role = fresh_user_data.get('role')
                        st.session_state.last_permission_refresh = time.time()
                        logger.info(f"Permissões atualizadas para o usuário {st.session_state.username}")
        st.session_state.last_activity = time.time()

    def change_password(self, username: str, old_password: str, new_password: str) -> bool:
        if not self._validate_password_strength(new_password):
            st.error("A nova senha deve ter 8+ caracteres, com maiúscula, minúscula e número.");
            return False
        user_df = self.db.get_docs("users", [("username", "==", username)])
        if user_df.empty: st.error("Usuário não encontrado."); return False
        user_data = user_df.iloc[0]
        if not self._check_password(user_data['password'], user_data['salt'], old_password):
            st.error("A senha antiga está incorreta.");
            return False
        new_hashed_pw, new_salt = self._hash_password(new_password)
        update_data = {"password": new_hashed_pw, "salt": new_salt}
        if self.db.update_doc("users", user_data['id'], update_data, username):
            st.success("Senha alterada com sucesso!");
            return True
        else:
            st.error("Não foi possível alterar a senha."); return False


# -----------------------------------------------------------------------------
# 3. UI / VIEWS (LÓGICA DE APRESENTAÇÃO)
# -----------------------------------------------------------------------------

def parse_brazilian_float(text: str) -> float:
    if not isinstance(text, str) or not text: return 0.0
    try:
        return float(text.replace('.', '').replace(',', '.'))
    except ValueError:
        st.error(f"Valor '{text}' inválido. Use o formato 1.234,56");
        raise


def format_brazilian_currency(value: float) -> str:
    if not isinstance(value, (int, float)): return "R$ 0,00"
    return f"R$ {value:_.2f}".replace('.', ',').replace('_', '.')


def to_excel(df: pd.DataFrame, title: str = "Relatório") -> bytes:
    output = io.BytesIO()
    df_copy = df.copy()
    for col in df_copy.columns:
        if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
            try:
                if hasattr(df_copy[col].dtype, 'tz') and df_copy[col].dtype.tz is not None:
                    df_copy[col] = df_copy[col].dt.tz_localize(None)
                elif df_copy[col].dtype == 'object':
                    df_copy[col] = pd.to_datetime(df_copy[col], errors='coerce')
                    if hasattr(df_copy[col].dtype, 'tz') and df_copy[col].dtype.tz is not None:
                        df_copy[col] = df_copy[col].dt.tz_localize(None)
            except Exception as e:
                logger.warning(f"Não foi possível processar timezone da coluna {col}: {e}");
                pass
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_copy.to_excel(writer, index=False, sheet_name=title)
        workbook, worksheet = writer.book, writer.sheets[title]
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col_num, col_name in enumerate(df_copy.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, alignment
            max_len = max(df_copy[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_len, 50)
        for row in range(2, len(df_copy) + 2):
            for col in range(1, len(df_copy.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border, cell.alignment = border, Alignment(horizontal='left', vertical='center')
                if 'valor' in df_copy.columns[col - 1].lower(): cell.number_format = 'R$ #,##0.00'
        status_col_name = next((col for col in ['status', 'status_demanda'] if col in df_copy.columns), None)
        if status_col_name:
            fills = {'green': PatternFill(start_color="C6EFCE", fill_type="solid"),
                     'red': PatternFill(start_color="FFC7CE", fill_type="solid"),
                     'yellow': PatternFill(start_color="FFEB9C", fill_type="solid"),
                     'blue': PatternFill(start_color="DDEBF7", fill_type="solid")}
            status_col_index = df_copy.columns.get_loc(status_col_name) + 1
            for row in range(2, len(df_copy) + 2):
                cell = worksheet.cell(row=row, column=status_col_index)
                if cell.value in ['Finalizado', 'Entregue', 'Fechada']:
                    cell.fill = fills['green']
                elif cell.value in ['Cancelado', 'Rejeitado']:
                    cell.fill = fills['red']
                elif cell.value in ['Em Processamento', 'Em Atendimento', 'Em Transporte', 'Pedido Gerado']:
                    cell.fill = fills['yellow']
                elif cell.value in ['Aberto', 'Aberta']:
                    cell.fill = fills['blue']
        worksheet.freeze_panes = 'A2';
        worksheet.auto_filter.ref = worksheet.dimensions
    return output.getvalue()


class ViewManager:
    def __init__(self, auth_service: AuthService, db_service: FirebaseService):
        self.auth, self.db = auth_service, db_service
        self._init_session_state()

    def _init_session_state(self):
        defaults = {'logged_in': False, 'username': "", 'role': "", 'page': "Login", 'user_data': {},
                    'confirm_delete': {}, 'edit_id': None,
                    'ignored_notifications': set(),  # <--- ADICIONE ISSO (conjunto para performance)
                    'edit_user_id': None, 'confirm_delete_user': {}, 'reset_password_for_user': {}, 'focus_item': None,
                    'view_history_id': None,
                    'generate_pedido_from_rc': None, 'confirm_restore': None, 'show_notifications': False,
                    'notifications_list': [],
                    'editing_comment': None, 'confirm_delete_comment': None, 'concluir_cadastro_id': None}
        for key, value in defaults.items():
            if key not in st.session_state: st.session_state[key] = value

    def _set_edit_state(self, collection: str, item_data: dict):
        st.session_state.edit_id = {'collection': collection, 'id': item_data['id'], 'data': item_data}

    def _set_delete_state(self, collection: str, item_id: str, item_desc: str):
        st.session_state.confirm_delete = {'collection': collection, 'id': item_id, 'desc': item_desc}

    def _set_history_state(self, collection: str, item_data: dict):
        st.session_state.view_history_id = {'collection': collection, 'id': item_data['id'], 'data': item_data}

    def _set_focus_state(self, collection: str, item_id: str):
        st.session_state.focus_item = {'collection': collection, 'id': item_id}

    def _set_generate_pedido_state(self, item_data: dict):
        st.session_state.generate_pedido_from_rc = item_data

    def _has_permission(self, permission: str) -> bool:
        if st.session_state.get("role") == "admin": return True
        return permission in st.session_state.user_data.get("permissions", [])

    def _validate_uploaded_file(self, uploaded_file, max_size_kb=750) -> tuple[bool, str]:
        allowed_mime_types = {
            'application/pdf': ['.pdf'], 'image/jpeg': ['.jpg', '.jpeg'], 'image/png': ['.png'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': ['.xlsx'],
            'application/vnd.ms-excel': ['.xls'], 'application/msword': ['.doc'],
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': ['.docx'],
            'text/plain': ['.txt']
        }
        if uploaded_file.size > max_size_kb * 1024: return False, f"Arquivo muito grande! Tamanho máximo: {max_size_kb}KB"
        if uploaded_file.size == 0: return False, "Arquivo está vazio"
        file_ext = os.path.splitext(uploaded_file.name)[1].lower()
        all_allowed_extensions = [ext for exts in allowed_mime_types.values() for ext in exts]
        if file_ext not in all_allowed_extensions: return False, f"Extensão não permitida: {file_ext}. Permitidas: {', '.join(all_allowed_extensions)}"
        if uploaded_file.type not in allowed_mime_types: return False, f"Tipo de arquivo não permitido: {uploaded_file.type}"
        if file_ext not in allowed_mime_types.get(uploaded_file.type,
                                                  []): return False, f"Extensão {file_ext} não corresponde ao tipo {uploaded_file.type}"
        file_bytes = uploaded_file.getvalue()
        if len(file_bytes) > 0:
            if file_ext == '.pdf' and not file_bytes.startswith(b'%PDF'): return False, "Arquivo não é um PDF válido"
            if file_ext == '.png' and not file_bytes.startswith(
                b'\x89PNG\r\n\x1a\n'): return False, "Arquivo não é um PNG válido"
            if file_ext in ['.jpg', '.jpeg'] and not file_bytes.startswith(
                b'\xff\xd8\xff'): return False, "Arquivo não é um JPEG válido"
        return True, "Arquivo válido"

    def run(self):
        if not st.session_state.logged_in:
            self.render_login_page()
        else:
            self.auth.check_session_timeout(); self.render_main_app()

    def render_login_page(self):
        _, col2, _ = st.columns([1, 2, 1])
        with col2:
            st.markdown(
                """<div style="text-align: center; margin-bottom: 2rem;"><span style="font-family: sans-serif; font-size: 4rem; font-weight: 900; color: var(--text-color);">ATIBAIA</span><span style="font-family: sans-serif; font-size: 4rem; font-weight: 900; color: #00AEEF;">💧</span><div style="font-family: sans-serif; font-size: 2.5rem; color: #00AEEF; letter-spacing: 0.1rem; margin-top: -1rem;">SANEAMENTO</div></div>""",
                unsafe_allow_html=True)
            if st.session_state.page == "Login":
                self._render_login_form()
                if st.button("Não tem conta? Registre-se"): st.session_state.page = "Registro"; st.rerun()
            else:
                self._render_registration_form()
                if st.button("Já tem conta? Faça login"): st.session_state.page = "Login"; st.rerun()

    def _render_login_form(self):
        st.title("🔐 Login do Sistema")
        with st.form("login_form"):
            username, password = st.text_input("Nome de Usuário"), st.text_input("Senha", type="password")
            if st.form_submit_button("Entrar", type="primary"): self.auth.login_user(username, password)

    def _render_registration_form(self):
        st.title("📝 Registro de Novo Usuário")
        with st.form("registration_form"):
            username, email, password = st.text_input("Nome de Usuário"), st.text_input("E-mail"), st.text_input(
                "Senha", type="password")
            is_gestor = st.checkbox("Sou um gestor (requer aprovação do admin)")
            if st.form_submit_button("Registrar", type="primary"): self.auth.register_user(username, email, password,is_gestor)

    def render_edit_modal(self):
        """
        Modal completo de edição com:
        - Aba 1: Formulários específicos por tipo (Demanda, RC, Pedido)
        - Aba 2: Gestão de Anexos (Upload/Download) e Chat Completo (Editar/Excluir)
        """
        edit_info = st.session_state.edit_id
        collection = edit_info['collection']
        doc_id = edit_info['id']

        # Tenta buscar dados frescos do banco
        current_data = self.db.get_doc(collection, doc_id)

        # Se falhar (documento deletado por outro), usa o cache ou avisa
        if not current_data:
            st.error("Documento não encontrado ou excluído.")
            if st.button("Fechar"):
                st.session_state.edit_id = None
                st.rerun()
            return

        # --- CABEÇALHO DO MODAL ---
        desc_titulo = current_data.get('descricao_necessidade') or current_data.get('descricao') or "Item"
        st.markdown(f"### ✏️ Editando: {desc_titulo[:60]}...")

        # Cria as abas
        tab_dados, tab_extras = st.tabs(["📄 Dados Principais", "📎 Anexos e Comentários"])

        # =====================================================================
        # ABA 1: FORMULÁRIOS DE EDIÇÃO
        # =====================================================================
        with tab_dados:
            with st.form("edit_form"):
                update_data = {}

                # --- FORMULÁRIO: DEMANDAS ---
                if collection == "demandas":
                    descricao = st.text_area("Descrição Completa", value=current_data.get('descricao_necessidade', ''),
                                             height=150)

                    c1, c2 = st.columns(2)
                    cats = ["Facilities/Eletromecânica", "Manutenção de rede", "Tratamento", "Tratamento (Laboratório)"]
                    current_cat = current_data.get('categoria', cats[0])
                    # Proteção se a categoria salva não estiver na lista
                    idx_cat = cats.index(current_cat) if current_cat in cats else 0
                    categoria = c1.selectbox("Categoria", cats, index=idx_cat)

                    prio_opts = ["Baixa", "Média", "Alta"]
                    current_prio = current_data.get('prioridade', "Média")
                    idx_prio = prio_opts.index(current_prio) if current_prio in prio_opts else 1
                    prioridade = c2.selectbox("Prioridade", prio_opts, index=idx_prio)

                    c3, c4 = st.columns(2)
                    status_opts = ["Aberta", "Em Atendimento", "Pendente", "Cancelado", "Concluída"]
                    current_status = current_data.get('status_demanda', "Aberta")
                    idx_stat = status_opts.index(current_status) if current_status in status_opts else 0
                    status = c3.selectbox("Status", status_opts, index=idx_stat)

                    tipo_opts = ["Material", "Serviço"]
                    current_tipo = current_data.get('tipo', "Material")
                    idx_tipo = tipo_opts.index(current_tipo) if current_tipo in tipo_opts else 0
                    tipo = c4.selectbox("Tipo", tipo_opts, index=idx_tipo)

                    update_data = {
                        "descricao_necessidade": descricao,
                        "categoria": categoria,
                        "prioridade": prioridade,
                        "status_demanda": status,
                        "tipo": tipo
                    }

                # --- FORMULÁRIO: REQUISIÇÕES ---
                elif collection == "requisicoes":
                    val_atual = current_data.get('valor', 0.0)
                    val_str = f"{val_atual:.2f}".replace('.', ',')

                    c1, c2 = st.columns(2)
                    numero_rc = c1.text_input("Número RC", value=current_data.get('numero_rc', ''))
                    valor_input = c2.text_input("Valor (R$)", value=val_str)

                    status_opts = ["Aberto", "Pedido Gerado", "Cancelado"]
                    current_status = current_data.get('status', "Aberto")
                    idx_stat = status_opts.index(current_status) if current_status in status_opts else 0
                    status = st.selectbox("Status", status_opts, index=idx_stat)

                    try:
                        val_float = parse_brazilian_float(valor_input)
                        update_data = {"numero_rc": numero_rc, "valor": val_float, "status": status}
                    except:
                        st.error("Valor inválido")

                # --- FORMULÁRIO: PEDIDOS ---
                elif collection == "pedidos":
                    val_atual = current_data.get('valor', 0.0)
                    val_str = f"{val_atual:.2f}".replace('.', ',')

                    c1, c2 = st.columns(2)
                    numero_pedido = c1.text_input("Número Pedido", value=current_data.get('numero_pedido', ''))
                    valor_input = c2.text_input("Valor (R$)", value=val_str)

                    c3, c4 = st.columns(2)
                    status_opts = ["Em Processamento", "Em Transporte", "Entregue", "Cancelado"]
                    current_status = current_data.get('status', "Em Processamento")
                    idx_stat = status_opts.index(current_status) if current_status in status_opts else 0
                    status = c3.selectbox("Status", status_opts, index=idx_stat)

                    # Data de Entrega
                    data_entrega_atual = current_data.get('data_entrega')
                    if isinstance(data_entrega_atual, str):
                        try:
                            data_entrega_atual = datetime.fromisoformat(data_entrega_atual)
                        except:
                            data_entrega_atual = None
                    elif isinstance(data_entrega_atual, datetime):
                        data_entrega_atual = data_entrega_atual.date()

                    data_entrega = c4.date_input("Previsão de Entrega", value=data_entrega_atual)

                    try:
                        val_float = parse_brazilian_float(valor_input)
                        update_data = {
                            "numero_pedido": numero_pedido,
                            "valor": val_float,
                            "status": status,
                            "data_entrega": datetime.combine(data_entrega,
                                                             datetime.min.time()) if data_entrega else None
                        }
                    except:
                        st.error("Valor inválido")

                # --- FORMULÁRIO: CADASTROS ---
                elif collection == "solicitacoes_cadastro":
                    descricao = st.text_area("Descrição", value=current_data.get('descricao', ''))
                    update_data = {"descricao": descricao}

                else:
                    st.info("Edição básica não configurada para esta coleção.")

                st.divider()

                # BOTÕES DE AÇÃO DO FORMULÁRIO
                col_b1, col_b2 = st.columns([1, 1])
                if col_b1.form_submit_button("💾 Salvar Alterações", type="primary"):
                    if update_data:
                        update_data['updated_at'] = datetime.now()
                        if self.db.update_doc(collection, doc_id, update_data, st.session_state.username):
                            self.db.log_action("Item Edited", st.session_state.username,
                                               {"id": doc_id, "col": collection})
                            st.toast("✅ Salvo com sucesso!")
                            time.sleep(1)
                            st.session_state.edit_id = None
                            st.rerun()

                if col_b2.form_submit_button("Cancelar"):
                    st.session_state.edit_id = None
                    st.rerun()

        # =====================================================================
        # ABA 2: ANEXOS E COMENTÁRIOS (ESTILO PLANNER)
        # =====================================================================
        with tab_extras:
            col_files, col_chat = st.columns([1, 1])

            # --- COLUNA DA ESQUERDA: ARQUIVOS ---
            with col_files:
                st.subheader("📂 Arquivos")

                # Lista anexos (Array 'attachments' + legado 'anexo')
                attachments = current_data.get('attachments', [])
                old_anexo = current_data.get('anexo')

                # Compatibilidade: Se existir anexo antigo e não estiver na lista nova, adiciona visualmente
                display_attachments = attachments.copy()
                if old_anexo and isinstance(old_anexo, dict):
                    # Verifica duplicidade simples pelo nome
                    if not any(a.get('file_name') == old_anexo.get('file_name') for a in attachments):
                        display_attachments.insert(0, old_anexo)

                if not display_attachments:
                    st.caption("Nenhum arquivo anexado.")
                else:
                    for idx, file in enumerate(display_attachments):
                        with st.container(border=True):
                            c_icon, c_name, c_down, c_del = st.columns([0.15, 0.55, 0.15, 0.15])

                            # Ícone
                            f_name = file.get('file_name', 'Arquivo')
                            ext = os.path.splitext(f_name)[1].lower()
                            icon = "📄"
                            if ext in ['.pdf']:
                                icon = "📕"
                            elif ext in ['.xls', '.xlsx', '.csv']:
                                icon = "📊"
                            elif ext in ['.doc', '.docx']:
                                icon = "📝"
                            elif ext in ['.jpg', '.png', '.jpeg']:
                                icon = "🖼️"
                            c_icon.write(icon)

                            # Nome
                            c_name.write(f"**{f_name[:20]}...**" if len(f_name) > 20 else f"**{f_name}**")

                            # Download
                            try:
                                b64_data = file.get('b64_data')
                                if b64_data:
                                    file_bytes = base64.b64decode(b64_data)
                                    c_down.download_button("⬇️", file_bytes, file_name=f_name, key=f"dl_{doc_id}_{idx}")
                            except:
                                c_down.error("Err")

                            # Excluir (Chama método do Service)
                            if c_del.button("🗑️", key=f"rm_{doc_id}_{idx}"):
                                # Verifica se é anexo novo ou legado
                                if file in attachments:
                                    self.db.remove_attachment(collection, doc_id, file)
                                elif file == old_anexo:
                                    self.db.update_doc(collection, doc_id, {"anexo": firestore.DELETE_FIELD},
                                                       st.session_state.username)
                                st.rerun()

                st.divider()
                st.write("**Adicionar novo arquivo:**")
                new_file = st.file_uploader("Upload", type=['pdf', 'docx', 'xlsx', 'jpg', 'png', 'txt'],
                                            key=f"up_{doc_id}", label_visibility="collapsed")

                if new_file:
                    if st.button("Enviar Arquivo", key=f"btn_up_{doc_id}"):
                        is_valid, msg = self._validate_uploaded_file(new_file)
                        if is_valid:
                            b64 = base64.b64encode(new_file.getvalue()).decode('utf-8')
                            file_data = {
                                "file_name": new_file.name,
                                "content_type": new_file.type,
                                "b64_data": b64,
                                "uploaded_at": datetime.now().isoformat(),
                                "uploaded_by": st.session_state.username
                            }
                            # Chama método helper do FirebaseService
                            self.db.add_attachment(collection, doc_id, file_data)
                            st.toast("Arquivo anexado!")
                            st.rerun()
                        else:
                            st.error(msg)

            # --- COLUNA DA DIREITA: CHAT COMPLETO ---
            with col_chat:
                st.subheader("💬 Comentários")

                # Scroll Container
                msg_container = st.container(height=400, border=True)

                using_sub = current_data.get('using_comment_subcollection', False)

                # Busca comentários
                if using_sub:
                    comentarios = self.db.get_comments_from_subcollection(collection, doc_id)
                else:
                    comentarios = current_data.get('comentarios', [])
                    if not isinstance(comentarios, list): comentarios = []

                with msg_container:
                    if not comentarios:
                        st.info("Nenhum comentário.")

                    # Ordena: Antigo -> Novo
                    sorted_comments = sorted(comentarios, key=lambda x: x.get('timestamp', datetime.min) if isinstance(
                        x.get('timestamp'), datetime) else datetime.min)

                    for c in sorted_comments:
                        c_id = c.get('id')
                        is_me = c.get('username') == st.session_state.username
                        can_edit = is_me or st.session_state.role == 'admin'
                        avatar = "👤" if not is_me else "😎"

                        ts = c.get('timestamp')
                        if isinstance(ts, str): ts = datetime.fromisoformat(ts)
                        time_str = ts.strftime('%d/%m %H:%M') if ts else ""

                        with st.chat_message(c.get('username'), avatar=avatar):
                            # Header
                            h_col, act_col = st.columns([0.8, 0.2])
                            h_col.markdown(
                                f"**{c.get('username')}** <span style='font-size:0.7em; color:gray'>{time_str}</span>",
                                unsafe_allow_html=True)

                            # Botões (Editar/Excluir)
                            if can_edit:
                                with act_col:
                                    b1, b2 = st.columns(2)
                                    if b1.button("✏️", key=f"ed_{c_id}", help="Editar"):
                                        st.session_state.editing_comment = c_id
                                        st.rerun()
                                    if b2.button("🗑️", key=f"dl_{c_id}", help="Excluir"):
                                        if using_sub:
                                            self.db.db.collection(collection).document(doc_id).collection(
                                                'comments').document(c_id).delete()
                                        else:
                                            new_list = [x for x in comentarios if x.get('id') != c_id]
                                            self.db.update_doc(collection, doc_id, {"comentarios": new_list},
                                                               st.session_state.username)
                                        st.rerun()

                            # Modo Edição vs Leitura
                            if st.session_state.get('editing_comment') == c_id:
                                edit_txt = st.text_area("Editar", value=c.get('text'), key=f"et_{c_id}")
                                s_col, c_col = st.columns(2)
                                if s_col.button("Salvar", key=f"sv_{c_id}"):
                                    if using_sub:
                                        self.db.db.collection(collection).document(doc_id).collection(
                                            'comments').document(c_id).update(
                                            {"text": edit_txt, "edited_at": datetime.now()})
                                    else:
                                        for item in comentarios:
                                            if item['id'] == c_id:
                                                item['text'] = edit_txt
                                                item['edited_at'] = datetime.now()
                                                break
                                        self.db.update_doc(collection, doc_id, {"comentarios": comentarios},
                                                           st.session_state.username)
                                    del st.session_state.editing_comment
                                    st.rerun()
                                if c_col.button("Cancelar", key=f"cn_{c_id}"):
                                    del st.session_state.editing_comment
                                    st.rerun()
                            else:
                                # Leitura com quebra de linha corrigida
                                clean_text = self._format_comment_text(c.get('text', ''),
                                                                       get_cached_docs(self.db, "users"))
                                st.markdown(clean_text)
                                if 'edited_at' in c: st.caption("(editado)")

                # Input Novo Comentário
                new_msg = st.chat_input("Escreva um comentário...")
                if new_msg:
                    c_data = {
                        "id": str(uuid.uuid4()),
                        "username": st.session_state.username,
                        "timestamp": datetime.now(),
                        "text": new_msg
                    }

                    # Notificações @
                    all_users = get_cached_docs(self.db, "users")
                    valid_u = all_users['username'].tolist() if not all_users.empty else []
                    for u in set(re.findall(r'@(\w+)', new_msg)):
                        if u in valid_u and u != st.session_state.username:
                            self._create_mention_notification(u, st.session_state.username, collection, doc_id)

                    # Salvar
                    if using_sub or len(comentarios) >= 20:
                        if not using_sub:
                            self.db.migrate_comments_to_subcollection(collection, doc_id)
                        self.db.add_comment_to_subcollection(collection, doc_id, c_data)
                    else:
                        self.db.update_doc(collection, doc_id, {"comentarios": comentarios + [c_data]},
                                           st.session_state.username)
                    st.rerun()

    def render_data_row(self, row: pd.Series, collection: str, **kwargs):
        """Renderiza uma linha de dados (Card) nas listas de Demandas, RCs e Pedidos."""
        key, role = f"{collection}_{row['id']}", st.session_state.role

        with st.container(border=True):
            # 1. Cabeçalho do Card (Varia conforme o tipo)
            if collection == 'demandas':
                assigned_str = f" | **Atribuído a:** `{row.get('assigned_to')}`" if row.get('assigned_to') else ""
                title = f"Demanda: {row.get('descricao_necessidade', '')} (Tipo: {row.get('tipo', 'N/A')} | Cat: {row.get('categoria', 'N/A')})"
                st.markdown(
                    f"**{title}**\n\n**Status:** `{row.get('status_demanda', 'N/A')}` | **Criado por:** `{row.get('solicitante_demanda', 'N/A')}` em `{row.get('created_at').strftime('%d/%m/%Y')}`{assigned_str}")

            elif collection == 'requisicoes':
                title = f"RC: {row.get('numero_rc', 'S/N')} | Valor: {format_brazilian_currency(row.get('valor', 0))}"
                st.markdown(
                    f"**{title}**\n\n**Status:** `{row.get('status', 'N/A')}` | **Criado por:** `{row.get('solicitante', 'N/A')}` em `{row.get('created_at').strftime('%d/%m/%Y')}`")

            else:  # Pedidos
                title = f"Pedido: {row.get('numero_pedido', 'S/N')} | Valor: {format_brazilian_currency(row.get('valor', 0))}"
                st.markdown(
                    f"**{title}**\n\n**Status:** `{row.get('status', 'N/A')}` | **Criado por:** `{row.get('solicitante', 'N/A')}` em `{row.get('created_at').strftime('%d/%m/%Y')}`")

            # 2. Informações de Vínculo (Ex: Mostrar qual demanda gerou a RC)
            if collection in ['requisicoes', 'pedidos']:
                # Tenta achar o ID da demanda original
                demanda_id = row.get('demanda_id')
                if collection == 'pedidos':
                    # Se for pedido, precisa achar a RC primeiro para pegar o demanda_id dela
                    all_rcs = kwargs.get('all_rcs', pd.DataFrame())
                    if not all_rcs.empty:
                        rc_info = all_rcs[all_rcs['id'] == row.get('requisicao_id')]
                        if not rc_info.empty:
                            demanda_id = rc_info.iloc[0].get('demanda_id')

                if demanda_id:
                    demandas = kwargs.get('all_demandas', pd.DataFrame())
                    if not demandas.empty:
                        demanda_info = demandas[demandas['id'] == demanda_id]
                        if not demanda_info.empty:
                            with st.expander("Ver Descrição da Demanda Original"):
                                st.info(demanda_info.iloc[0]['descricao_necessidade'])

            # 3. Botões de Ação
            cols = st.columns([1, 1, 1, 2, 5])

            # Botão Editar
            is_author = (row.get('solicitante_demanda') or row.get('solicitante')) == st.session_state.username
            if is_author or self._has_permission("pode_excluir"):
                cols[0].button("✏️", key=f"edit_{key}", help="Editar", on_click=self._set_edit_state,
                               args=(collection, row.to_dict()))

            # Botão Excluir
            if self._has_permission("pode_excluir"):
                cols[1].button("🗑️", key=f"del_{key}", help="Excluir", on_click=self._set_delete_state,
                               args=(collection, row['id'], title))

            # Botão Histórico
            cols[2].button("📜", key=f"hist_{key}", help="Ver Histórico", on_click=self._set_history_state,
                           args=(collection, row.to_dict()))

            # Botões de Fluxo (Ir para Pedido / Gerar Pedido)
            if collection == 'demandas':
                all_rcs = kwargs.get('all_rcs')
                all_pedidos = kwargs.get('all_pedidos')

                # Verifica se tem RC ligada
                linked_rc = pd.DataFrame()
                if all_rcs is not None and not all_rcs.empty:
                    linked_rc = all_rcs[all_rcs['demanda_id'] == row['id']]

                if not linked_rc.empty:
                    rc_id = linked_rc.iloc[0]['id']
                    # Verifica se tem Pedido ligado à RC
                    linked_pedido = pd.DataFrame()
                    if all_pedidos is not None and not all_pedidos.empty:
                        linked_pedido = all_pedidos[all_pedidos['requisicao_id'] == rc_id]

                    if not linked_pedido.empty:
                        cols[3].button("🚚 Ver Pedido", key=f"goto_ped_{key}", on_click=self._set_focus_state,
                                       args=('pedidos', linked_pedido.iloc[0]['id']))
                    else:
                        cols[3].button("🛒 Ver RC", key=f"goto_rc_{key}", on_click=self._set_focus_state,
                                       args=('requisicoes', rc_id))

            if collection == "requisicoes" and row.get('status') == "Aberto" and role in ['admin', 'user']:
                cols[3].button("📦 Gerar Pedido", key=f"gen_ped_{key}", type="primary",
                               on_click=self._set_generate_pedido_state, args=(row.to_dict(),))

            # Lógica de Confirmação de Exclusão (Aparece se clicou na lixeira)
            if st.session_state.confirm_delete.get('id') == row['id']:
                st.warning(f"Excluir '{st.session_state.confirm_delete['desc']}'?")
                c1, c2, _ = st.columns([1, 1, 8])
                if c1.button("Sim, excluir", key=f"conf_del_{key}", type="primary"):
                    self.db.delete_doc(collection, row['id'])
                    self.db.log_action(f"{collection[:-1].capitalize()} Deleted", st.session_state.username,
                                       {"doc_id": row['id'], "description": title})
                    st.session_state.confirm_delete = {}
                    st.rerun()
                if c2.button("Cancelar", key=f"canc_del_{key}"):
                    st.session_state.confirm_delete = {}
                    st.rerun()
            # 4. Seção de Comentários
            with st.expander("💬 Comentários"):
                self._render_comments_section(row, collection, **kwargs)

    def _render_comments_section(self, row, collection, **kwargs):
        """
        Renderiza a seção de comentários com layout simplificado para garantir
        que os botões apareçam mesmo dentro de expanders.
        """
        doc_id = row['id']
        using_sub = row.get('using_comment_subcollection', False)

        # 1. CARREGA COMENTÁRIOS
        if using_sub:
            comentarios = self.db.get_comments_from_subcollection(collection, doc_id, limit=MAX_COMMENTS_DISPLAY)
        else:
            comentarios = row.get('comentarios', [])
            if not isinstance(comentarios, list): comentarios = []

            # Migração automática
            if len(comentarios) >= MAX_COMMENTS_IN_DOCUMENT:
                with st.spinner("Otimizando comentários..."):
                    if self.db.migrate_comments_to_subcollection(collection, doc_id):
                        st.rerun()

        # 2. EXIBE A LISTA
        if not comentarios:
            st.caption("Nenhum comentário ainda.")
        else:
            # Ordena: Antigo -> Novo
            sorted_comments = sorted(comentarios, key=lambda x: x.get('timestamp', datetime.min) if isinstance(x.get('timestamp'), datetime) else datetime.min)

            for c in sorted_comments:
                c_id = c.get('id')
                # Verifica permissão
                is_me = c.get('username') == st.session_state.username
                can_edit = is_me or st.session_state.role == 'admin'

                avatar = "👤" if not is_me else "😎"

                # Formata data
                ts = c.get('timestamp')
                if isinstance(ts, str): ts = datetime.fromisoformat(ts)
                time_str = ts.strftime('%d/%m %H:%M') if ts else ""

                with st.chat_message(c.get('username'), avatar=avatar):

                    # --- LAYOUT ROBUSTO (SEM COLUNAS ANINHADAS) ---
                    # Divide a linha do cabeçalho em 3: [Texto] [Editar] [Excluir]
                    if can_edit:
                        # Se pode editar, reserva espaço para os botões
                        c_head, c_edit, c_del = st.columns([0.8, 0.1, 0.1])
                    else:
                        # Se não pode editar, o texto ocupa tudo
                        c_head = st.columns([1])[0]
                        c_edit, c_del = None, None

                    # Renderiza Nome e Data
                    with c_head:
                        st.markdown(f"**{c.get('username')}** <span style='font-size:0.75em; color:gray'>({time_str})</span>", unsafe_allow_html=True)

                    # Renderiza Botões (direto na coluna principal, sem criar novas colunas dentro)
                    if can_edit:
                        with c_edit:
                            if st.button("✏️", key=f"ed_{c_id}_{doc_id}", help="Editar"):
                                st.session_state.editing_comment = c_id
                                st.rerun()
                        with c_del:
                            if st.button("🗑️", key=f"dl_{c_id}_{doc_id}", help="Excluir"):
                                if using_sub:
                                    self.db.db.collection(collection).document(doc_id).collection('comments').document(c_id).delete()
                                else:
                                    new_list = [x for x in comentarios if x.get('id') != c_id]
                                    self.db.update_doc(collection, doc_id, {"comentarios": new_list}, st.session_state.username)
                                st.rerun()

                    # --- CONTEÚDO (Visualização ou Edição) ---
                    if st.session_state.get('editing_comment') == c_id:
                        # MODO EDIÇÃO
                        edit_txt = st.text_area("Editar:", value=c.get('text'), key=f"txt_{c_id}_{doc_id}")
                        col_s, col_c = st.columns([1, 1])

                        if col_s.button("✅ Salvar", key=f"sav_{c_id}_{doc_id}"):
                            if using_sub:
                                self.db.db.collection(collection).document(doc_id).collection('comments').document(c_id).update({
                                    "text": edit_txt, "edited_at": datetime.now()
                                })
                            else:
                                for item in comentarios:
                                    if item.get('id') == c_id:
                                        item['text'] = edit_txt
                                        item['edited_at'] = datetime.now()
                                        break
                                self.db.update_doc(collection, doc_id, {"comentarios": comentarios}, st.session_state.username)

                            del st.session_state.editing_comment
                            st.rerun()

                        if col_c.button("Cancelar", key=f"can_{c_id}_{doc_id}"):
                            del st.session_state.editing_comment
                            st.rerun()
                    else:
                        # MODO LEITURA (Usando sua função de formatação)
                        safe_txt = self._format_comment_text(c.get('text', ''), kwargs.get('all_users', pd.DataFrame()))
                        st.markdown(safe_txt)
                        if 'edited_at' in c:
                            st.caption("(editado)")

        # 3. CAMPO PARA NOVO COMENTÁRIO
        new_comment = st.text_area("Adicionar comentário", key=f"add_c_{doc_id}", height=68, label_visibility="collapsed", placeholder="Escreva um comentário...")
        if st.button("Enviar", key=f"snd_c_{doc_id}"):
            if new_comment:
                c_data = {
                    "id": str(uuid.uuid4()),
                    "username": st.session_state.username,
                    "timestamp": datetime.now(),
                    "text": new_comment
                }

                # Notificações
                all_users_df = kwargs.get('all_users', pd.DataFrame())
                valid_users = all_users_df['username'].tolist() if not all_users_df.empty else []
                for u in set(re.findall(r'@(\w+)', new_comment)):
                    if u in valid_users and u != st.session_state.username:
                        self._create_mention_notification(u, st.session_state.username, collection, doc_id)

                # Salvar
                if using_sub or len(comentarios) >= 20:
                    if not using_sub:
                        self.db.migrate_comments_to_subcollection(collection, doc_id)
                        using_sub = True
                    self.db.add_comment_to_subcollection(collection, doc_id, c_data)
                else:
                    self.db.update_doc(collection, doc_id, {"comentarios": comentarios + [c_data]}, st.session_state.username)

                st.rerun()

    def render_main_app(self):
        self.render_sidebar()

        # Se estiver em modo de edição (modal de tela cheia), mostra o editor
        if st.session_state.edit_id:
            self.render_edit_modal()
        # Se estiver focado em um item específico (visualização detalhada), mostra o foco
        elif st.session_state.focus_item:
            self.render_focused_view()
        else:
            # Cabeçalho Principal da Aplicação
            col1, col2 = st.columns([0.8, 0.2])
            col1.title("🚀 Sistema de Controle de Compras")
            with col2:
                self.render_notification_bell()

            # --- DEFINIÇÃO DAS ABAS ---
            # Aqui adicionamos "📋 Quadro" como a segunda aba
            tabs = ["📊 Dashboard", "📋 Quadro", "📝 Demandas", "🛒 Requisições", "🚚 Pedidos", "📜 Controle de Cadastros"]

            # Aba extra visível apenas para Admin
            if st.session_state.role == 'admin':
                tabs.append("🛡️ Registros de Atividades")

            # Cria os componentes visuais das abas
            selected_tabs = st.tabs(tabs)

            # --- CONTEÚDO DAS ABAS ---

            # 1. Dashboard
            with selected_tabs[0]:
                self.render_dashboard()

            # 2. Quadro Kanban (Novo recurso estilo Planner)
            with selected_tabs[1]:
                self.render_planner_tab()

            # 3. Demandas (Lista)
            with selected_tabs[2]:
                self.render_demandas()

            # 4. Requisições
            with selected_tabs[3]:
                self.render_requisicoes()

            # 5. Pedidos
            with selected_tabs[4]:
                self.render_pedidos()

            # 6. Controle de Cadastros
            with selected_tabs[5]:
                self.render_controle_cadastro()

            # 7. Logs (Apenas Admin)
            if st.session_state.role == 'admin':
                with selected_tabs[6]:  # Índice 6 pois inserimos o Quadro antes
                    self.render_logs_tab()

        # --- MODAIS FLUTUANTES (DIALOGS) ---
        # Renderizados fora da estrutura principal para aparecerem por cima
        if st.session_state.view_history_id:
            self.render_history_modal()

        if st.session_state.generate_pedido_from_rc:
            self.render_generate_pedido_modal()

        if st.session_state.get('show_notifications', False):
            self.render_notifications_modal()

        if st.session_state.get('concluir_cadastro_id'):
            self.render_concluir_cadastro_modal()

    def render_sidebar(self):
        with st.sidebar:
            st.write(f"👤 **{st.session_state.username}** ({st.session_state.role})")
            with st.expander("Meu Perfil", expanded=True):
                with st.form("change_password_form", clear_on_submit=True):
                    st.subheader("Alterar Senha")
                    old_p, new_p, conf_p = st.text_input("Senha Antiga", type="password"), st.text_input("Nova Senha",
                                                                                                         type="password",
                                                                                                         help="Mínimo 8 caracteres, com maiúscula, minúscula e número."), st.text_input(
                        "Confirmar Nova Senha", type="password")
                    if st.form_submit_button("Alterar Senha", type="primary"):
                        if new_p != conf_p:
                            st.error("As novas senhas não coincidem.")
                        else:
                            if self.auth.change_password(st.session_state.username, old_p, new_p):
                                self.db.log_action("Password Changed", st.session_state.username);
                                time.sleep(2);
                                st.rerun()
            if st.button("Logout", use_container_width=True):
                self.db.log_action("User Logout", st.session_state.username)
                for key in list(st.session_state.keys()): del st.session_state[key]
                st.rerun()
            st.divider()
            if st.session_state.role == 'admin': self.render_admin_panel()

    def render_admin_panel(self):
        st.header("⚙️ Administração")
        with st.expander("Gerenciar Usuários", expanded=True):
            if st.session_state.edit_user_id:
                self._render_edit_user_form()
            else:
                self._render_user_lists()
        st.divider();
        st.subheader("Backup e Restauro Local")
        if st.download_button(label="📥 Baixar Backup Local", data=self._generate_backup_data(),
                              file_name=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                              mime="application/json", use_container_width=True, type="primary"):
            self.db.log_action("Backup Downloaded", st.session_state.username)
        uploaded_file = st.file_uploader("Restaurar a partir de arquivo (.json)", type="json",
                                         help="ATENÇÃO: Restaurar substituirá todos os dados existentes.")
        if uploaded_file:
            if st.button("Restaurar Backup"): st.session_state.confirm_restore = uploaded_file; st.rerun()
        if st.session_state.get('confirm_restore'):
            st.error(f"Restaurar '{st.session_state.confirm_restore.name}'? Dados atuais serão perdidos.")
            rc1, rc2, _ = st.columns([1, 1, 3])
            if rc1.button("Sim, restaurar", key="conf_restore_l", type="primary"):
                self.db.log_action("Backup Restored", st.session_state.username,
                                   {"file_name": st.session_state.confirm_restore.name})
                if self.db.restore_from_backup_data(json.load(st.session_state.confirm_restore)):
                    st.success("Backup restaurado!");
                    del st.session_state.confirm_restore;
                    time.sleep(2);
                    st.rerun()
            if rc2.button("Cancelar", key="canc_restore_l"): del st.session_state.confirm_restore; st.rerun()

    def _render_user_lists(self):
        # OTIMIZAÇÃO: Busca todos os usuários do cache de uma vez só
        all_users = get_cached_docs(self.db, "users")

        # Filtra PENDENTES na memória
        pending_users = all_users[all_users["status"] == "pending"] if not all_users.empty else pd.DataFrame()

        if not pending_users.empty:
            st.subheader("Aprovações Pendentes")
            for _, user in pending_users.iterrows():
                c1, c2, c3 = st.columns([2, 1, 1])
                c1.write(f"{user['username']} ({user['role']})")
                if c2.button("✅", key=f"a_{user['id']}", help="Aprovar"):
                    self.db.update_doc("users", user['id'], {"status": "active"}, st.session_state.username)
                    self.db.log_action("User Approved", st.session_state.username, {"approved_user": user['username']})
                    st.rerun()
                if c3.button("🗑️", key=f"r_{user['id']}", help="Rejeitar"):
                    self.db.delete_doc("users", user['id'])
                    self.db.log_action("User Rejected", st.session_state.username, {"rejected_user": user['username']})
                    st.rerun()
            st.divider()

        # Filtra ATIVOS na memória
        st.subheader("Usuários Ativos")
        active_users = all_users[all_users["status"] == "active"] if not all_users.empty else pd.DataFrame()

        for _, user in active_users.iterrows():
            is_self = user['username'] == st.session_state.username
            c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
            c1.write(f"**{user['username']}** ({user.get('email', 'sem e-mail')}) - `{user['role']}`")
            if c2.button("✏️", key=f"edit_user_{user['id']}", help="Editar"):
                st.session_state.edit_user_id = user['id']
                st.rerun()
            if c3.button("🔑", key=f"reset_pw_{user['id']}", help="Redefinir Senha", disabled=is_self):
                st.session_state.reset_password_for_user = {'id': user['id'], 'username': user['username']}
                st.rerun()
            if c4.button("🗑️", key=f"del_user_{user['id']}", help="Excluir", disabled=is_self):
                st.session_state.confirm_delete_user = {'id': user['id'], 'username': user['username']}
                st.rerun()

    def _render_edit_user_form(self):
        user_data = self.db.get_doc("users", st.session_state.edit_user_id)
        st.subheader(f"Editando Usuário: {user_data['username']}")
        AVAILABLE_PERMISSIONS = {"pode_excluir": "Pode Excluir Itens"}
        with st.form("edit_user_form"):
            email = st.text_input("E-mail", value=user_data.get('email', ''))
            role = st.selectbox("Cargo", ["user", "gestor", "admin"],
                                index=["user", "gestor", "admin"].index(user_data.get('role', 'user')))
            departments = ["N/A", "Operacional", "Manutenção", "Administrativo", "Financeiro"]
            current_dept = user_data.get('department', 'N/A')
            dept_index = departments.index(current_dept) if current_dept in departments else 0
            department = st.selectbox("Departamento", departments, index=dept_index)
            st.divider();
            st.subheader("Permissões Especiais")
            current_perms = user_data.get('permissions', [])
            selected_perms = st.multiselect("Permissões", options=list(AVAILABLE_PERMISSIONS.keys()),
                                            default=current_perms, format_func=lambda p: AVAILABLE_PERMISSIONS[p])
            c1, c2 = st.columns(2)
            if c1.form_submit_button("Salvar", type="primary"):
                if email != user_data.get('email', ''):
                    existing_user = self.db.get_docs("users", [("email", "==", email)])
                    if not existing_user.empty:
                        st.error(f"O e-mail '{email}' já está em uso por outro usuário.");
                        return
                try:
                    User(username=user_data['username'], email=email, role=role,
                         status=user_data.get('status', 'active'), department=department, permissions=selected_perms)
                    update_data = {"email": email, "role": role, "department": department,
                                   "permissions": selected_perms}
                    if self.db.update_doc("users", user_data['id'], update_data, st.session_state.username):
                        self.db.log_action("User Edited", st.session_state.username,
                                           {"target_user": user_data['username'], "changes": update_data})
                        st.success("Usuário atualizado!");
                        st.session_state.edit_user_id = None;
                        time.sleep(1);
                        st.rerun()
                except ValidationError as e:
                    st.error(f"E-mail inválido: {e.errors()[0]['msg']}")
            if c2.form_submit_button("Cancelar"): st.session_state.edit_user_id = None; st.rerun()

    def render_notification_bell(self):
        all_notifications = []

        # Lógica de Admin (mantida)
        if st.session_state.role == 'admin':
            all_users = get_cached_docs(self.db, "users")
            if not all_users.empty:
                pending_users = all_users[all_users['status'] == 'pending']
                for _, user in pending_users.iterrows():
                    notif_id = f"admin_approval_{user['id']}"
                    # FILTRO NOVO: Só adiciona se não estiver na lista de ignorados
                    if notif_id not in st.session_state.ignored_notifications:
                        all_notifications.append({
                            "id": notif_id,
                            "message": f"Aprovação pendente: {user['username']}",
                            "type": "admin_approval"
                        })

        # Notificações normais do Cache
        all_notifs_df = get_cached_docs(self.db, "notifications")
        if not all_notifs_df.empty:
            user_notifications_df = all_notifs_df[
                (all_notifs_df['username'] == st.session_state.username) &
                (all_notifs_df['read'] == False)
                ]
            for _, notif in user_notifications_df.iterrows():
                # FILTRO NOVO: Verifica se já marcamos como lida nesta sessão
                if notif['id'] not in st.session_state.ignored_notifications:
                    all_notifications.append(notif.to_dict())

        st.session_state.notifications_list = all_notifications
        num_notifications = len(all_notifications)
        label = f"🔔 ({num_notifications})" if num_notifications > 0 else "🔔"

        # Se clicar, abre o modal
        if st.button(label, help="Notificações"):
            st.session_state.show_notifications = not st.session_state.get('show_notifications', False)
            # Apenas rerun simples, sem limpar cache
            st.rerun()

    def _render_paginated_rows(self, df: pd.DataFrame, render_function, key_suffix: str, **kwargs):
        if df.empty:
            st.info("Nenhum dado encontrado.");
            return
        try:
            df_hash = hashlib.md5(pd.util.hash_pandas_object(df.index, index=True).values).hexdigest()
        except:
            df_hash = str(len(df))
        df_hash_key, items_key, page_key = f"df_hash_{key_suffix}", f"items_{key_suffix}", f"page_{key_suffix}"
        if st.session_state.get(df_hash_key) != df_hash:
            st.session_state[page_key] = 1;
            st.session_state[df_hash_key] = df_hash

        def reset_page_callback():
            st.session_state[page_key] = 1

        items_per_page = st.selectbox("Itens por página", [5, 10, 20, 50], key=items_key, index=1,
                                      on_change=reset_page_callback)
        total_pages = max(1, (len(df) - 1) // items_per_page + 1)
        current_page = st.session_state.get(page_key, 1)
        if current_page > total_pages or current_page < 1: st.session_state[page_key] = 1; current_page = 1
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            if st.button("⬅️ Anterior", key=f"prev_{key_suffix}", disabled=(current_page <= 1)):
                st.session_state[page_key] = max(1, current_page - 1);
                st.rerun()
        with col2:
            st.markdown(
                f"<div style='text-align: center; padding: 5px;'>Página <strong>{current_page}</strong> de <strong>{total_pages}</strong> ({len(df)} itens no total)</div>",
                unsafe_allow_html=True)
        with col3:
            if st.button("Próxima ➡️", key=f"next_{key_suffix}", disabled=(current_page >= total_pages)):
                st.session_state[page_key] = min(total_pages, current_page + 1);
                st.rerun()
        start_idx = (current_page - 1) * items_per_page
        end_idx = min(start_idx + items_per_page, len(df))
        for _, row in df.iloc[start_idx:end_idx].iterrows(): render_function(row, **kwargs)

    def render_focused_view(self):
        focus_info = st.session_state.focus_item
        collection, doc_id = focus_info['collection'], focus_info['id']
        item_map = {"demandas": "Demanda", "requisicoes": "Requisição", "pedidos": "Pedido"}
        item_type = item_map.get(collection, collection[:-1]).capitalize()

        st.subheader(f"Navegação: {item_type}s > Visualizando {item_type}")
        if st.button("⬅️ Voltar para a lista"):
            st.session_state.focus_item = None
            st.rerun()

        # Busca o item específico (aqui usamos get_doc direto para garantir o dado mais fresco possível do item atual)
        doc_data = self.db.get_doc(collection, doc_id)

        if doc_data:
            row = pd.Series(doc_data)

            # OTIMIZAÇÃO: As tabelas auxiliares vêm do CACHE
            all_users = get_cached_docs(self.db, "users")
            all_demandas = get_cached_docs(self.db, "demandas") if collection != 'demandas' else None
            all_rcs = get_cached_docs(self.db, "requisicoes") if collection == 'pedidos' else None

            self.render_data_row(row, collection=collection, all_demandas=all_demandas, all_rcs=all_rcs,
                                 all_users=all_users)
        else:
            st.error("Item não encontrado.")

    def render_dashboard(self):
        st.header("📊 Dashboard de Métricas")

        # 1. Carrega TODOS os dados brutos do banco
        df_demandas = get_cached_docs(self.db, "demandas")
        df_rc = get_cached_docs(self.db, "requisicoes")
        df_pedidos = get_cached_docs(self.db, "pedidos")

        # 2. Lógica para identificar os Anos Disponíveis
        anos_disponiveis = set()
        current_year = datetime.now().year

        # Verifica anos nas demandas
        if not df_demandas.empty and 'created_at' in df_demandas.columns:
            anos_disponiveis.update(df_demandas['created_at'].dt.year.unique())

        # Verifica anos nas requisições
        if not df_rc.empty and 'created_at' in df_rc.columns:
            anos_disponiveis.update(df_rc['created_at'].dt.year.unique())

        # Adiciona o ano atual (para garantir que ele apareça mesmo se não houver dados ainda)
        anos_disponiveis.add(current_year)

        # Transforma em lista ordenada
        lista_anos = sorted(list(anos_disponiveis), reverse=True)

        # 3. Cria o Seletor de Ano na tela (Padrão: Ano Atual)
        c_filtro, _ = st.columns([1, 3])
        with c_filtro:
            ano_selecionado = st.selectbox("📅 Selecione o Ano de Referência", lista_anos, index=0)

        # 4. Filtra os DataFrames com base no ano selecionado
        # Se o dataframe não estiver vazio, filtra pelo ano. Se estiver vazio, mantém vazio.
        df_demandas_ano = df_demandas[
            df_demandas['created_at'].dt.year == ano_selecionado] if not df_demandas.empty else df_demandas
        df_rc_ano = df_rc[df_rc['created_at'].dt.year == ano_selecionado] if not df_rc.empty else df_rc
        df_pedidos_ano = df_pedidos[
            df_pedidos['created_at'].dt.year == ano_selecionado] if not df_pedidos.empty else df_pedidos

        # --- A PARTIR DAQUI, USAMOS APENAS OS DATAFRAMES FILTRADOS (_ano) ---

        self.render_tutorial()

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric(f"Demandas ({ano_selecionado})", f"{len(df_demandas_ano)} 📝")
        c2.metric(f"RCs ({ano_selecionado})", f"{len(df_rc_ano)} 🛒")
        c3.metric(f"Pedidos ({ano_selecionado})", f"{len(df_pedidos_ano)} 🚚")

        valor_total_rc = df_rc_ano['valor'].sum() if not df_rc_ano.empty else 0
        c4.metric("Valor em RCs", format_brazilian_currency(valor_total_rc))

        valor_total_pedidos = df_pedidos_ano['valor'].sum() if not df_pedidos_ano.empty else 0
        c5.metric("Valor em Pedidos", format_brazilian_currency(valor_total_pedidos))

        st.divider()

        c1, c2 = st.columns(2)
        with c1:
            st.subheader(f"Status das Demandas - {ano_selecionado}")
            if not df_demandas_ano.empty:
                status_counts = df_demandas_ano['status_demanda'].value_counts().reset_index()
                # Ajuste para versão mais recente do Pandas/Plotly (nomes das colunas podem variar)
                status_counts.columns = ['status_demanda', 'count']

                fig = px.bar(status_counts, x='status_demanda', y='count',
                             title=f"Distribuição de Status em {ano_selecionado}",
                             text_auto=True,
                             color='status_demanda',
                             labels={'status_demanda': 'Status', 'count': 'Quantidade'})
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info(f"Nenhuma demanda registrada em {ano_selecionado}.")

        with c2:
            st.subheader(f"Demandas por Categoria - {ano_selecionado}")
            if not df_demandas_ano.empty:
                cat_counts = df_demandas_ano['categoria'].value_counts().reset_index()
                cat_counts.columns = ['categoria', 'count']

                fig = px.pie(cat_counts, names='categoria', values='count',
                             title=f"Distribuição por Categoria em {ano_selecionado}",
                             hole=.3,
                             labels={'categoria': 'Categoria', 'count': 'Quantidade'})
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info(f"Nenhuma categoria registrada em {ano_selecionado}.")

    def render_tutorial(self):
        with st.expander("💡 Mini-Tutorial: Como Usar o Sistema", expanded=False):
            st.markdown("""
            **1. Crie uma Demanda (📝):** Na aba **Demandas**, clique em **"➕ Adicionar Nova Demanda"**.
            **2. Crie uma Requisição (🛒):** Na aba **Requisições**, clique em **"➕ Adicionar Nova Requisição"** e selecione a demanda.
            **3. Gere um Pedido (🚚):** Na lista de **Requisições**, clique em **"📦 Gerar Pedido"**.
            **Colaboração:** Use **"💬 Comentários"** e mencione usuários com `@nome`.
            """)

    def _reset_page_number(self, page_key_suffix):
        if f"page_{page_key_suffix}" in st.session_state: st.session_state[f"page_{page_key_suffix}"] = 1

    def render_advanced_filters(self, df: pd.DataFrame, key_prefix: str) -> pd.DataFrame:
        with st.expander("🔍 Filtros Avançados"):
            if df.empty: return df
            filtered_df = df.copy();
            cols = st.columns(3)
            with cols[0]:
                search_term = st.text_input("Buscar por texto...", key=f"search_{key_prefix}",
                                            on_change=self._reset_page_number, args=(key_prefix,))
                if search_term:
                    text_cols = filtered_df.select_dtypes(include=['object']).columns.tolist()
                    filtered_df = filtered_df[filtered_df[text_cols].apply(
                        lambda row: row.astype(str).str.contains(search_term, case=False).any(), axis=1)]
            with cols[1]:
                filter_cols_map = {'status_demanda': 'Status da Demanda', 'status': 'Status', 'categoria': 'Categoria',
                                   'tipo': 'Tipo', 'solicitante_demanda': 'Solicitante', 'solicitante': 'Solicitante',
                                   'assigned_to': 'Atribuído a'}
                for col, label in filter_cols_map.items():
                    if col in filtered_df.columns:
                        options = sorted(filtered_df[col].dropna().unique())
                        if options:
                            selected = st.multiselect(label, options, key=f"filter_{col}_{key_prefix}",
                                                      on_change=self._reset_page_number, args=(key_prefix,))
                            if selected: filtered_df = filtered_df[filtered_df[col].isin(selected)]
            with cols[2]:
                if 'created_at' in filtered_df.columns:
                    start_date = st.date_input("De:", value=None, key=f"start_date_{key_prefix}",
                                               on_change=self._reset_page_number, args=(key_prefix,))
                    end_date = st.date_input("Até:", value=None, key=f"end_date_{key_prefix}",
                                             on_change=self._reset_page_number, args=(key_prefix,))
                    date_series = pd.to_datetime(filtered_df['created_at']).dt.date
                    if start_date: filtered_df = filtered_df[date_series >= start_date]
                    if end_date: filtered_df = filtered_df[date_series <= end_date]
            return filtered_df

    def render_demandas(self):
        st.header("📝 Demandas de Compras")
        if st.session_state.role in ['admin', 'user', 'gestor']:
            with st.expander("➕ Adicionar Nova Demanda"):
                with st.form("demanda_form", clear_on_submit=True):
                    # OTIMIZAÇÃO: Usa lista de usuários do cache
                    user_list = get_cached_users_list(self.db)
                    descricao, assigned_to = st.text_area("Descrição da Necessidade"), st.selectbox("Atribuir para:",
                                                                                                    user_list)
                    c1, c2 = st.columns(2)
                    tipo = c1.selectbox("Tipo", ["Material", "Serviço"], index=None, placeholder="Selecione o tipo...")
                    categorias_fixas = ["Facilities/Eletromecânica", "Manutenção de rede", "Tratamento",
                                        "Tratamento (Laboratório)"]
                    categoria = c2.selectbox("Categoria", categorias_fixas, index=None,
                                             placeholder="Selecione a categoria...")
                    uploaded_file = st.file_uploader("Anexo (Opcional, máx 750KB)",
                                                     type=['pdf', 'jpg', 'jpeg', 'png', 'xlsx', 'xls', 'doc', 'docx',
                                                           'txt'])

                    if st.form_submit_button("Registrar Demanda", type="primary"):
                        if not all([descricao, categoria, tipo]):
                            st.error("Preencha todos os campos obrigatórios (Descrição, Tipo e Categoria).")
                            return
                        with st.spinner("Registrando demanda..."):
                            anexo_data_dict = None
                            if uploaded_file:
                                is_valid, validation_message = self._validate_uploaded_file(uploaded_file)
                                if not is_valid:
                                    st.error(validation_message)
                                    return
                                b64_data = base64.b64encode(uploaded_file.getvalue()).decode('utf-8')
                                anexo_data_dict = {"file_name": uploaded_file.name, "content_type": uploaded_file.type,
                                                   "b64_data": b64_data, "file_size": uploaded_file.size}
                            try:
                                demanda = Demanda(solicitante_demanda=st.session_state.username,
                                                  descricao_necessidade=descricao, tipo=tipo, categoria=categoria,
                                                  anexo=anexo_data_dict,
                                                  assigned_to=assigned_to if assigned_to != "Ninguém" else None)
                                demanda_data = demanda.model_dump()
                                demanda_data['historico'] = [
                                    f"Criado por {st.session_state.username} em {datetime.now().strftime('%d/%m/%Y %H:%M')}"]

                                new_demanda_id = self.db.add_doc("demandas", demanda_data)
                                if new_demanda_id:
                                    self.db.log_action("Demanda Created", st.session_state.username,
                                                       {"doc_id": new_demanda_id, "description": descricao,
                                                        "assigned_to": assigned_to})
                                    if assigned_to and assigned_to != "Ninguém":
                                        self._create_assignment_notification(assigned_user=assigned_to,
                                                                             author=st.session_state.username,
                                                                             collection="demandas",
                                                                             doc_id=new_demanda_id,
                                                                             description=descricao)
                                    st.toast("✅ Demanda registrada!", icon="✅")
                                    time.sleep(1)
                                    st.rerun()
                            except ValidationError as e:
                                st.error(f"Erro de validação: {e}")

            with st.expander("➕ Adicionar Múltiplas Demandas (via Planilha)"):
                self._render_bulk_upload_section()

        st.header("Demandas Registradas")
        # OTIMIZAÇÃO: Usa get_cached_docs para tudo
        df_demandas = get_cached_docs(self.db, "demandas")
        df_rcs = get_cached_docs(self.db, "requisicoes")
        df_pedidos = get_cached_docs(self.db, "pedidos")
        df_users = get_cached_docs(self.db, "users")

        filtered_demandas = self.render_advanced_filters(df_demandas, "demandas")
        if not filtered_demandas.empty:
            st.download_button(label="📥 Exportar para Excel", data=to_excel(filtered_demandas, "Demandas"),
                               file_name="demandas_exportadas.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="export_demandas")

        st.divider()
        self._render_paginated_rows(filtered_demandas, self.render_data_row, "demandas", collection="demandas",
                                    all_rcs=df_rcs, all_pedidos=df_pedidos, all_users=df_users)

    def _render_bulk_upload_section(self):
        st.info(
            "Faça o upload de uma planilha Excel (.xlsx) com as colunas: `descricao_necessidade`, `tipo`, `categoria`.")
        df_modelo = pd.DataFrame({"descricao_necessidade": ["Exemplo: Compra de 10 capacetes"], "tipo": ["Material"],
                                  "categoria": ["Facilities/Eletromecânica"]})
        output = io.BytesIO();
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_modelo.to_excel(writer, index=False, sheet_name='Modelo')
        st.download_button(label="📥 Baixar Planilha Modelo", data=output.getvalue(), file_name="modelo_demandas.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        uploaded_file = st.file_uploader("Selecione a planilha", type="xlsx")
        if uploaded_file and st.button("Processar Planilha", type="primary"):
            try:
                df = pd.read_excel(uploaded_file);
                df.columns = [col.lower().replace(" ", "_") for col in df.columns]
                required_columns = ["descricao_necessidade", "tipo", "categoria"]
                if not all(col in df.columns for col in required_columns): st.error(
                    f"A planilha deve conter as colunas: {', '.join(required_columns)}"); return
                success_count, error_list, total_rows = 0, [], len(df)
                progress_bar = st.progress(0, text="Processando demandas...")
                tipos_validos, categorias_validas = ["Material", "Serviço"], ["Facilities/Eletromecânica",
                                                                              "Manutenção de rede", "Tratamento",
                                                                              "Tratamento (Laboratório)"]
                for index, row in df.iterrows():
                    try:
                        descricao, tipo, categoria = row['descricao_necessidade'], row['tipo'], row['categoria']
                        if not (descricao and tipo and categoria): raise ValueError("Dados obrigatórios em branco.")
                        if tipo not in tipos_validos: raise ValueError(f"Tipo '{tipo}' inválido.")
                        if categoria not in categorias_validas: raise ValueError(f"Categoria '{categoria}' inválida.")
                        demanda = Demanda(solicitante_demanda=st.session_state.username,
                                          descricao_necessidade=str(descricao), tipo=str(tipo),
                                          categoria=str(categoria))
                        demanda_data = demanda.model_dump();
                        demanda_data['historico'] = [
                            f"Criado por {st.session_state.username} via upload em {datetime.now().strftime('%d/%m/%Y %H:%M')}"]
                        if self.db.add_doc("demandas", demanda_data):
                            success_count += 1
                        else:
                            raise Exception("Falha ao salvar no banco de dados.")
                    except Exception as e:
                        error_list.append(f"Linha {index + 2}: {e} | Dados: {row.to_dict()}")
                    progress_bar.progress((index + 1) / total_rows, text=f"Processando {index + 1}/{total_rows}")
                self.db.log_action("Bulk Demanda Upload", st.session_state.username,
                                   {"file_name": uploaded_file.name, "success_count": success_count,
                                    "error_count": len(error_list)})
                st.success(f"{success_count} de {total_rows} demandas registradas!")
                if error_list: st.error("Algumas linhas não puderam ser processadas:"); [st.write(e) for e in
                                                                                         error_list]
                time.sleep(3);
                st.rerun()
            except Exception as e:
                st.error(f"Erro ao processar o arquivo: {e}")

    def _prepare_df_for_export_rc(self, df_rc: pd.DataFrame, df_demandas: pd.DataFrame) -> pd.DataFrame:
        if df_rc.empty or df_demandas.empty or 'demanda_id' not in df_rc.columns: return df_rc
        df_demandas_subset = df_demandas[['id', 'descricao_necessidade']].copy()
        merged_df = pd.merge(df_rc, df_demandas_subset, left_on='demanda_id', right_on='id', how='left',
                             suffixes=('', '_demanda'))
        merged_df['descricao_necessidade'] = merged_df['descricao_necessidade'].fillna('Demanda não encontrada')
        merged_df = merged_df.rename(columns={'descricao_necessidade': 'Descrição da Demanda'})
        cols_to_drop = ['demanda_id', 'id_demanda']
        return merged_df.drop(columns=[col for col in cols_to_drop if col in merged_df.columns], errors='ignore')

    def render_requisicoes(self):
        st.header("🛒 Requisições de Compra (RCs)")
        if st.session_state.role in ['admin', 'user']:
            with st.expander("➕ Adicionar Nova Requisição"):
                st.info("Crie uma nova Requisição de Compra a partir de uma demanda que ainda não foi atendida.")
                st.subheader("Passo 1: Selecione a Demanda")

                # OTIMIZAÇÃO: Carrega todas demandas do CACHE e filtra 'Aberta' na memória
                all_demandas = get_cached_docs(self.db, "demandas")
                if not all_demandas.empty:
                    df_demandas_abertas = all_demandas[all_demandas['status_demanda'] == "Aberta"]
                else:
                    df_demandas_abertas = pd.DataFrame()

                demanda_options = {"Selecione uma Demanda": None,
                                   **{f"ID: ...{r['id'][-6:]} - {r['descricao_necessidade'][:40]}...": r['id'] for _, r
                                      in df_demandas_abertas.iterrows()}}
                selected_demanda_key = st.selectbox("Vincular à Demanda", list(demanda_options.keys()),
                                                    label_visibility="collapsed")
                selected_demanda_id = demanda_options.get(selected_demanda_key)

                if selected_demanda_id:
                    details = df_demandas_abertas[df_demandas_abertas['id'] == selected_demanda_id].iloc[0]
                    with st.container(border=True):
                        st.markdown("##### Detalhes da Demanda Selecionada")
                        st.text_area("Descrição da Necessidade", value=details['descricao_necessidade'], height=150,
                                     disabled=True)
                        c1, c2, c3 = st.columns(3)
                        c1.markdown(f"**Tipo:**\n\n`{details.get('tipo', 'N/A')}`")
                        c2.markdown(f"**Categoria:**\n\n`{details['categoria']}`")
                        c3.markdown(f"**Solicitante:**\n\n`{details['solicitante_demanda']}`")
                        anexo_info = details.get('anexo')
                        if anexo_info and isinstance(anexo_info, dict) and 'b64_data' in anexo_info:
                            try:
                                file_bytes = base64.b64decode(anexo_info['b64_data'])
                                st.download_button(label=f"📥 Baixar anexo original: {anexo_info['file_name']}",
                                                   data=file_bytes, file_name=anexo_info['file_name'],
                                                   mime=anexo_info.get('content_type', 'application/octet-stream'))
                            except Exception as e:
                                st.error(f"Não foi possível carregar o anexo: {e}")

                    st.subheader("Passo 2: Detalhes da Requisição")
                    with st.form("requisicao_form_details", clear_on_submit=True):
                        valor_str, numero_rc = st.text_input("Valor (R$)", placeholder="Ex: 1.234,56"), st.text_input(
                            "Número da RC (opcional)")
                        if st.form_submit_button("Registrar Requisição", type="primary"):
                            try:
                                valor = parse_brazilian_float(valor_str)
                                if valor <= 0:
                                    st.error("O valor deve ser maior que zero.")
                                    return
                                requisicao = Requisicao(solicitante=st.session_state.username,
                                                        demanda_id=selected_demanda_id, valor=valor,
                                                        numero_rc=numero_rc or None)
                                req_data = requisicao.model_dump()
                                req_data['historico'] = [
                                    f"Criado por {st.session_state.username} em {datetime.now().strftime('%d/%m/%Y %H:%M')}"]
                                update_demanda_data = {"status_demanda": "Em Atendimento", "updated_at": datetime.now()}
                                if self.db.add_and_update_atomically("requisicoes", req_data, "demandas",
                                                                     selected_demanda_id, update_demanda_data):
                                    self.db.log_action("Requisicao Created", st.session_state.username,
                                                       {"demanda_id": selected_demanda_id, "valor": valor})
                                    st.toast("✅ Requisição registrada!", icon="✅")
                                    time.sleep(1)
                                    st.rerun()
                            except ValueError:
                                return
                            except Exception as e:
                                st.error(f"Erro ao registrar: {e}")

        st.header("Requisições Registradas")
        # OTIMIZAÇÃO: Usa cache
        df_rc = get_cached_docs(self.db, "requisicoes")
        df_demandas = get_cached_docs(self.db, "demandas")
        df_users = get_cached_docs(self.db, "users")

        filtered_rcs = self.render_advanced_filters(df_rc, "requisicoes")
        if not filtered_rcs.empty:
            df_for_export = self._prepare_df_for_export_rc(filtered_rcs, df_demandas)
            st.download_button(label="📥 Exportar para Excel", data=to_excel(df_for_export, "Requisições"),
                               file_name="requisicoes_exportadas.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="export_requisicoes")

        st.divider()
        self._render_paginated_rows(filtered_rcs, self.render_data_row, "rcs", collection="requisicoes",
                                    all_demandas=df_demandas, all_users=df_users)

    def _prepare_df_for_export_pedidos(self, df_pedidos: pd.DataFrame, df_rcs: pd.DataFrame,
                                       df_demandas: pd.DataFrame) -> pd.DataFrame:
        if df_pedidos.empty: return df_pedidos
        if df_rcs.empty or df_demandas.empty: df_pedidos[
            'Descrição da Demanda'] = 'Dados de origem indisponíveis'; return df_pedidos
        df_merged_rc = pd.merge(df_pedidos, df_rcs[['id', 'demanda_id']], left_on='requisicao_id', right_on='id',
                                how='left', suffixes=('', '_rc'))
        df_final = pd.merge(df_merged_rc, df_demandas[['id', 'descricao_necessidade']], left_on='demanda_id',
                            right_on='id', how='left', suffixes=('', '_demanda'))
        df_final['descricao_necessidade'] = df_final['descricao_necessidade'].fillna('Demanda original não encontrada')
        df_final = df_final.rename(columns={'descricao_necessidade': 'Descrição da Demanda'})
        cols_to_drop = ['requisicao_id', 'id_rc', 'demanda_id', 'id_demanda']
        return df_final.drop(columns=[col for col in cols_to_drop if col in df_final.columns], errors='ignore')

    def render_pedidos(self):
        st.header("🚚 Pedidos de Compra")
        # OTIMIZAÇÃO: Usa cache para todas as tabelas
        all_pedidos = get_cached_docs(self.db, "pedidos")
        all_rcs = get_cached_docs(self.db, "requisicoes")
        all_demandas = get_cached_docs(self.db, "demandas")
        all_users = get_cached_docs(self.db, "users")

        tabs = st.tabs(["⏳ Em Andamento", "✅ Entregues", "❌ Cancelados"])
        status_map = [['Em Processamento', 'Em Transporte'], ['Entregue'], ['Cancelado']]

        for tab, statuses in zip(tabs, status_map):
            with tab:
                # Filtra na memória
                df_tab_filtered = all_pedidos[
                    all_pedidos['status'].isin(statuses)] if not all_pedidos.empty else pd.DataFrame()
                final_filtered_pedidos = self.render_advanced_filters(df_tab_filtered, f"pedidos_{statuses[0]}")

                if not final_filtered_pedidos.empty:
                    df_for_export = self._prepare_df_for_export_pedidos(final_filtered_pedidos, all_rcs, all_demandas)
                    st.download_button("📥 Exportar para Excel", to_excel(df_for_export, f"Pedidos {statuses[0]}"),
                                       f'pedidos_{statuses[0].lower().replace(" ", "_")}.xlsx',
                                       key=f'btn_export_{statuses[0]}')

                st.divider()
                self._render_paginated_rows(final_filtered_pedidos, self.render_data_row, f"pedidos_{statuses[0]}",
                                            collection="pedidos", all_rcs=all_rcs, all_demandas=all_demandas,
                                            all_users=all_users)

    def render_controle_cadastro(self):
        st.header("📜 Controle de Cadastro de Itens")

        # --- PARTE NOVA (OTIMIZADA) ---

        # 1. Carrega TODAS as solicitações do cache de uma vez só
        # Em vez de ir ao banco duas vezes, pegamos tudo aqui.
        df_all_cadastros = get_cached_docs(self.db, "solicitacoes_cadastro")

        # 2. Carrega lista de usuários do cache (para passar pros métodos internos depois)
        all_users = get_cached_docs(self.db, "users")

        with st.expander("➕ Solicitar Novo Cadastro de Item"):
            with st.form("solicitacao_cadastro_form", clear_on_submit=True):
                descricao = st.text_area("Descrição do item a ser cadastrado")
                if st.form_submit_button("Enviar Solicitação", type="primary"):
                    if not descricao or len(descricao) < 5:
                        st.warning("A descrição é obrigatória e deve ter pelo menos 5 caracteres.")
                    else:
                        nova_solicitacao = SolicitacaoCadastro(solicitante=st.session_state.username,
                                                               descricao=descricao)
                        if self.db.add_doc("solicitacoes_cadastro", nova_solicitacao.model_dump()):
                            st.success("Solicitação de cadastro enviada!")
                            # Como salvamos algo novo, precisamos limpar o cache para ele aparecer na lista
                            clear_cache()
                            time.sleep(1)
                            st.rerun()

        st.divider()
        tab_cadastrando, tab_concluido = st.tabs(["⏳ Cadastrando", "✅ Concluído"])

        # --- FILTRAGEM NA MEMÓRIA ---

        with tab_cadastrando:
            # Aqui filtramos usando Pandas (memória) em vez de chamar self.db.get_docs
            if not df_all_cadastros.empty:
                # Pega apenas as linhas onde status == 'Cadastrando'
                df_cadastrando = df_all_cadastros[df_all_cadastros['status'] == "Cadastrando"]
            else:
                df_cadastrando = pd.DataFrame()

            if not df_cadastrando.empty:
                st.download_button("📥 Exportar para Excel", to_excel(df_cadastrando, "Cadastros Pendentes"),
                                   'cadastros_pendentes.xlsx')
                for _, item in df_cadastrando.iterrows():
                    self._render_solicitacao_cadastro_row(item, all_users)
            else:
                st.info("Nenhuma solicitação de cadastro pendente.")

        with tab_concluido:
            # Aqui filtramos novamente a mesma tabela carregada no início
            if not df_all_cadastros.empty:
                # Pega apenas as linhas onde status == 'Concluído'
                df_concluido = df_all_cadastros[df_all_cadastros['status'] == "Concluído"]
            else:
                df_concluido = pd.DataFrame()

            if not df_concluido.empty:
                st.download_button("📥 Exportar para Excel", to_excel(df_concluido, "Cadastros Concluídos"),
                                   'cadastros_concluidos.xlsx')
                for _, item in df_concluido.iterrows():
                    self._render_solicitacao_cadastro_row(item, all_users)
            else:
                st.info("Nenhum item concluído.")

    def _render_solicitacao_cadastro_row(self, item_row, all_users):
        """Renderiza uma única linha de solicitação de cadastro com ações."""
        collection = "solicitacoes_cadastro"
        item_id = item_row['id']
        key = f"{collection}_{item_id}"
        with st.container(border=True):
            if item_row['status'] == 'Concluído':
                st.write(
                    f"**Solicitante:** {item_row['solicitante']} | **Concluído em:** {item_row['updated_at'].strftime('%d/%m/%Y')}")
                st.caption(f"**Descrição Original:** {item_row['descricao']}")
                st.success(f"**Código do Item Cadastrado:** `{item_row['codigo_item_final']}`")
            else:
                st.write(f"**Solicitante:** {item_row['solicitante']} em {item_row['created_at'].strftime('%d/%m/%Y')}")
                st.info(item_row['descricao'])
            cols = st.columns([1, 1, 2, 6])
            with cols[0]:
                st.button("✏️", key=f"edit_{key}", help="Editar Descrição", on_click=self._set_edit_state,
                          args=(collection, item_row.to_dict()))
            with cols[1]:
                st.button("🗑️", key=f"del_{key}", help="Excluir Solicitação", on_click=self._set_delete_state,
                          args=(collection, item_id, item_row['descricao']))
            if item_row['status'] == 'Cadastrando':
                with cols[2]:
                    if st.button("Concluir Cadastro", key=f"concluir_{key}", type="primary"):
                        st.session_state.concluir_cadastro_id = item_row.to_dict()
                        st.rerun()
            if st.session_state.confirm_delete.get('id') == item_id:
                st.warning(f"Excluir solicitação '{st.session_state.confirm_delete['desc']}'?")
                c1, c2, _ = st.columns([1, 1, 8])
                if c1.button("Sim, excluir", key=f"conf_del_{key}", type="primary"):
                    self.db.delete_doc(collection, item_id)
                    self.db.log_action("Solicitação de Cadastro Excluída", st.session_state.username,
                                       {"doc_id": item_id})
                    st.session_state.confirm_delete = {};
                    st.rerun()
                if c2.button("Cancelar", key=f"canc_del_{key}"):
                    st.session_state.confirm_delete = {};
                    st.rerun()
            with st.expander("💬 Comentários"):
                self._render_comments_section(item_row, collection, all_users=all_users)

    @st.dialog("Concluir Cadastro de Item")
    def render_concluir_cadastro_modal(self):
        item_data = st.session_state.concluir_cadastro_id
        st.write("Descrição do item:")
        st.info(item_data['descricao'])
        codigo_final = st.text_input("Digite o código do item que foi cadastrado no sistema:")
        col1, col2 = st.columns(2)
        if col1.button("Salvar e Concluir", type="primary"):
            if not codigo_final:
                st.error("O código do item é obrigatório para concluir.")
            else:
                update_data = {"status": "Concluído", "codigo_item_final": codigo_final, "updated_at": datetime.now()}
                if self.db.update_doc("solicitacoes_cadastro", item_data['id'], update_data, st.session_state.username):
                    st.toast("Cadastro concluído com sucesso!")
                    st.session_state.concluir_cadastro_id = None
                    st.rerun()
        if col2.button("Cancelar"):
            st.session_state.concluir_cadastro_id = None
            st.rerun()

    def _format_comment_text(self, text: str, all_users_df: pd.DataFrame) -> str:
        """Formata o texto para corrigir quebras de linha e menções."""
        if not isinstance(text, str): return ""

        # 1. Força quebra de linha visual (Markdown precisa de dois espaços ou quebra dupla)
        # Transforma qualquer 'Enter' em 'Dois Enters' para garantir parágrafo
        clean_text = text.replace('\r\n', '\n').replace('\n', '\n\n')

        # 2. Escapa caracteres especiais que podem quebrar o visual (exceto negrito/italico basico)
        # Se quiser permitir formatação básica, remova esta parte ou ajuste
        # clean_text = clean_text.replace('*', '\\*').replace('_', '\\_')

        # 3. Formata Menções (@usuario fica em negrito)
        user_mentions = re.findall(r'@(\w+)', clean_text)
        if user_mentions:
            valid_usernames = all_users_df['username'].tolist() if not all_users_df.empty else []
            for username in set(user_mentions):
                if username in valid_usernames:
                    clean_text = clean_text.replace(f"@{username}", f"**@{username}**")

        return clean_text

    def _create_mention_notification(self, mentioned_user: str, author: str, collection: str, doc_id: str):
        item_type = {"demandas": "Demanda", "requisicoes": "Requisição", "pedidos": "Pedido"}.get(collection,
                                                                                                  collection[:-1])
        notification_data = {"username": mentioned_user, "author": author,
                             "message": f"{author} mencionou você nos comentários da {item_type}.",
                             "link": {"collection": collection, "id": doc_id}, "read": False,
                             "timestamp": datetime.now()}
        self.db.add_doc("notifications", notification_data)
        self.db.log_action("Mention Notification Sent", author, {"to_user": mentioned_user, "doc_id": doc_id})

    def _create_assignment_notification(self, assigned_user: str, author: str, collection: str, doc_id: str,
                                        description: str):
        """ Cria uma notificação quando um item é atribuído a um usuário. """
        if assigned_user == author:  # Não notificar a si mesmo
            return

        item_type = {"demandas": "Demanda", "requisicoes": "Requisição", "pedidos": "Pedido"}.get(collection,
                                                                                                  collection[:-1])

        # Limita a descrição para não poluir a notificação
        desc_short = (description[:40] + '...') if len(description) > 40 else description

        notification_data = {
            "username": assigned_user,
            "author": author,
            "message": f"{author} atribuiu a {item_type} '{desc_short}' a você.",
            "link": {"collection": collection, "id": doc_id},
            "read": False,
            "timestamp": datetime.now()
        }
        self.db.add_doc("notifications", notification_data)
        self.db.log_action("Assignment Notification Sent", author, {"to_user": assigned_user, "doc_id": doc_id})

    def _render_comments_section(self, row, collection, **kwargs):
        """
        Renderiza comentários com botões de ação VISÍVEIS e texto formatado.
        """
        doc_id = row['id']
        using_sub = row.get('using_comment_subcollection', False)

        # 1. CARREGA
        if using_sub:
            comentarios = self.db.get_comments_from_subcollection(collection, doc_id, limit=MAX_COMMENTS_DISPLAY)
        else:
            comentarios = row.get('comentarios', [])
            if not isinstance(comentarios, list): comentarios = []

            if len(comentarios) >= MAX_COMMENTS_IN_DOCUMENT:
                with st.spinner("Otimizando..."):
                    if self.db.migrate_comments_to_subcollection(collection, doc_id): st.rerun()

        # 2. LISTA
        if not comentarios:
            st.caption("Nenhum comentário.")
        else:
            # Ordena: Antigo -> Novo
            sorted_comments = sorted(comentarios,
                                     key=lambda x: x.get('timestamp', datetime.min) if isinstance(x.get('timestamp'),
                                                                                                  datetime) else datetime.min)

            for c in sorted_comments:
                c_id = c.get('id')
                is_me = c.get('username') == st.session_state.username
                can_edit = is_me or st.session_state.role == 'admin'
                avatar = "👤" if not is_me else "😎"

                # Data
                ts = c.get('timestamp')
                if isinstance(ts, str): ts = datetime.fromisoformat(ts)
                time_str = ts.strftime('%d/%m %H:%M') if ts else ""

                with st.chat_message(c.get('username'), avatar=avatar):

                    # --- MODO EDIÇÃO ---
                    if st.session_state.get('editing_comment') == c_id:
                        st.write(f"**Editando comentário de {time_str}:**")
                        edit_txt = st.text_area("Texto", value=c.get('text'), key=f"t_{c_id}_{doc_id}",
                                                label_visibility="collapsed")
                        b1, b2 = st.columns([1, 1])

                        if b1.button("💾 Salvar", key=f"sv_{c_id}_{doc_id}", type="primary", use_container_width=True):
                            if using_sub:
                                self.db.db.collection(collection).document(doc_id).collection('comments').document(
                                    c_id).update({"text": edit_txt, "edited_at": datetime.now()})
                            else:
                                for item in comentarios:
                                    if item.get('id') == c_id:
                                        item['text'] = edit_txt;
                                        item['edited_at'] = datetime.now();
                                        break
                                self.db.update_doc(collection, doc_id, {"comentarios": comentarios},
                                                   st.session_state.username)
                            del st.session_state.editing_comment
                            st.rerun()

                        if b2.button("❌ Cancelar", key=f"cn_{c_id}_{doc_id}", use_container_width=True):
                            del st.session_state.editing_comment
                            st.rerun()

                    # --- MODO VISUALIZAÇÃO ---
                    else:
                        # 1. Cabeçalho + Botões na mesma linha (Layout 80% / 20%)
                        # Usamos pesos inteiros grandes para evitar erro de float
                        cols = st.columns([8, 2])

                        with cols[0]:
                            st.markdown(
                                f"**{c.get('username')}** <span style='font-size:0.75em; color:gray'> • {time_str}</span>",
                                unsafe_allow_html=True)

                        # Botões aparecem aqui se tiver permissão
                        if can_edit:
                            with cols[1]:
                                # Container horizontal para agrupar os botões
                                sub_c1, sub_c2 = st.columns(2)
                                sub_c1.button("✏️", key=f"ed_{c_id}_{doc_id}", help="Editar",
                                              on_click=lambda id=c_id: st.session_state.update({'editing_comment': id}))
                                if sub_c2.button("🗑️", key=f"dl_{c_id}_{doc_id}", help="Excluir"):
                                    if using_sub:
                                        self.db.db.collection(collection).document(doc_id).collection(
                                            'comments').document(c_id).delete()
                                    else:
                                        new_list = [x for x in comentarios if x.get('id') != c_id]
                                        self.db.update_doc(collection, doc_id, {"comentarios": new_list},
                                                           st.session_state.username)
                                    st.rerun()

                        # 2. Texto do Comentário (Abaixo do cabeçalho)
                        safe_txt = self._format_comment_text(c.get('text', ''), kwargs.get('all_users', pd.DataFrame()))
                        st.markdown(safe_txt)

                        if 'edited_at' in c: st.caption(f"_(editado)_")

        # 3. NOVO COMENTÁRIO
        st.write("")  # Espaçamento
        new_comment = st.text_area("Novo Comentário", key=f"new_{doc_id}", height=68, label_visibility="collapsed",
                                   placeholder="Escreva aqui...")

        if st.button("Enviar Comentário", key=f"send_{doc_id}"):
            if new_comment:
                c_data = {"id": str(uuid.uuid4()), "username": st.session_state.username, "timestamp": datetime.now(),
                          "text": new_comment}

                # Menções
                all_users_df = kwargs.get('all_users', pd.DataFrame())
                valid_u = all_users_df['username'].tolist() if not all_users_df.empty else []
                for u in set(re.findall(r'@(\w+)', new_comment)):
                    if u in valid_u and u != st.session_state.username:
                        self._create_mention_notification(u, st.session_state.username, collection, doc_id)

                # Salvar
                if using_sub or len(comentarios) >= 20:
                    if not using_sub:
                        self.db.migrate_comments_to_subcollection(collection, doc_id)
                        using_sub = True
                    self.db.add_comment_to_subcollection(collection, doc_id, c_data)
                else:
                    self.db.update_doc(collection, doc_id, {"comentarios": comentarios + [c_data]},
                                       st.session_state.username)
                st.rerun()

    @st.dialog("Histórico de Alterações")
    def render_history_modal(self):
        info = st.session_state.view_history_id
        st.markdown(f"**ID:** `{info['id']}`")
        for entry in reversed(info['data'].get('historico', ["Nenhum histórico."])): st.info(entry)
        if st.button("Fechar"): st.session_state.view_history_id = None; st.rerun()

    @st.dialog("Gerar Pedido de Compra")
    def render_generate_pedido_modal(self):
        rc_data = st.session_state.generate_pedido_from_rc
        st.write(
            f"Gerando pedido para a RC: **{rc_data.get('numero_rc', 'S/N')}** | Valor Original: **{format_brazilian_currency(rc_data.get('valor', 0))}**")
        with st.form("generate_pedido_form"):
            numero_pedido = st.text_input("Número do Pedido",
                                          value=f"PED-{rc_data.get('numero_rc', rc_data['id'][-4:])}")
            valor_sugerido = f"{rc_data.get('valor', 0.0):_.2f}".replace('.', ',').replace('_', '.')
            valor_final_str = st.text_input("Valor Final do Pedido (R$)", value=valor_sugerido)
            email_notificacao_str = st.text_area("E-mail para Notificação (opcional)")
            anexo_email_file = st.file_uploader("Anexo para E-mail (opcional)",
                                                type=['pdf', 'jpg', 'jpeg', 'png', 'xlsx', 'xls', 'doc', 'docx', 'txt'])
            if st.form_submit_button("Confirmar", type="primary"):
                with st.spinner("Gerando pedido..."):
                    anexo_email_data = None
                    if anexo_email_file:
                        is_valid, validation_message = self._validate_uploaded_file(anexo_email_file)
                        if not is_valid: st.error(validation_message); return
                        b64_data = base64.b64encode(anexo_email_file.getvalue()).decode('utf-8')
                        anexo_email_data = {"file_name": anexo_email_file.name, "content_type": anexo_email_file.type,
                                            "b64_data": b64_data}
                    try:
                        valor_final_pedido = parse_brazilian_float(valor_final_str)
                        if valor_final_pedido <= 0:
                            st.error("O valor final do pedido deve ser maior que zero.");
                            return

                        emails_para_notificar_lista = []
                        if email_notificacao_str:
                            # 1. Cria a lista de e-mails a partir da string de entrada
                            lista_bruta = re.split(r'[;, \n]+', email_notificacao_str)
                            emails_para_notificar_lista = [email.strip() for email in lista_bruta if email.strip()]

                        # 2. Converte a lista em uma string única, separada por vírgulas
                        emails_para_notificar_string = ", ".join(emails_para_notificar_lista)

                        # 3. Cria o pedido passando a STRING, que é o que o modelo e a Cloud Function esperam agora
                        pedido = Pedido(requisicao_id=rc_data['id'],
                                        solicitante=rc_data['solicitante'],
                                        valor=valor_final_pedido,
                                        numero_pedido=numero_pedido,
                                        email_notificacao=emails_para_notificar_string or None,
                                        anexo_email=anexo_email_data)

                        pedido_data = pedido.model_dump();
                        pedido_data['historico'] = [
                            f"Criado por {st.session_state.username} em {datetime.now().strftime('%d/%m/%Y %H:%M')}"]
                        update_rc_data = {"status": "Pedido Gerado", "updated_at": datetime.now()}
                        if self.db.add_and_update_atomically("pedidos", pedido_data, "requisicoes", rc_data['id'],
                                                             update_rc_data):
                            self.db.log_action("Pedido Created", st.session_state.username,
                                               {"rc_id": rc_data['id'], "numero_pedido": numero_pedido,
                                                "valor": valor_final_pedido})
                            st.toast("Pedido gerado!", icon="🚀");
                            st.session_state.generate_pedido_from_rc = None;
                            time.sleep(1);
                            st.rerun()
                    except ValidationError as e:
                        st.error(f"Erro de validação, verifique os campos. Detalhes: {e}")
                    except ValueError:
                        pass
        if st.button("Cancelar"): st.session_state.generate_pedido_from_rc = None; st.rerun()

    def _convert_firestore_types(self, obj):
        if hasattr(obj, 'isoformat'): return obj.isoformat()
        if isinstance(obj, bytes): return base64.b64encode(obj).decode('utf-8')
        if isinstance(obj, dict): return {key: self._convert_firestore_types(value) for key, value in obj.items()}
        if isinstance(obj, list): return [self._convert_firestore_types(element) for element in obj]
        return obj

    def _generate_backup_data(self) -> bytes:
        try:
            backup_data = {}
            collections_to_backup = ["users", "demandas", "requisicoes", "pedidos", "audit_logs", "notifications"]
            for col in collections_to_backup:
                # OTIMIZAÇÃO: Usa get_cached_docs para gerar backup (mais rápido e econômico)
                docs_df = get_cached_docs(self.db, col)
                if docs_df.empty:
                    backup_data[col] = []
                    continue

                # Converte DataFrame para lista de dicionários
                records = docs_df.to_dict(orient='records')
                # Processa tipos complexos (datas, bytes)
                processed_records = [self._convert_firestore_types(rec) for rec in records]
                backup_data[col] = processed_records

            return json.dumps(backup_data, ensure_ascii=False, indent=4).encode('utf-8')
        except Exception as e:
            logger.error(f"Falha ao gerar dados de backup: {e}", exc_info=True)
            st.error(f"Erro ao gerar backup: {e}")
            return b""

    def render_planner_tab(self):
        st.header("📋 Quadro de Tarefas")

        # 1. Carrega dados
        df_demandas = get_cached_docs(self.db, "demandas")
        if df_demandas.empty:
            st.info("Nenhuma demanda para exibir.")
            return

        # ---------------------------------------------------------
        # A. BARRA DE FERRAMENTAS (FILTROS)
        # ---------------------------------------------------------
        with st.expander("🔎 Filtros e Visualização", expanded=True):
            c1, c2, c3 = st.columns(3)

            # 1. Lista de Usuários (Tratamento seguro)
            # Garante que não hajam valores nulos ou erros de tipo na lista
            unique_users = sorted(df_demandas['assigned_to'].dropna().astype(str).unique().tolist())
            users = ["Todos"] + unique_users

            # 2. Lógica Anti-Loop: Define o padrão apenas na primeira vez
            key_filter_user = "planner_filter_user_key"

            if key_filter_user not in st.session_state:
                # Se for a primeira vez, tenta selecionar o usuário logado
                if st.session_state.username in users:
                    st.session_state[key_filter_user] = st.session_state.username
                else:
                    st.session_state[key_filter_user] = "Todos"

            # 3. Selectboxes (Criação das variáveis f_user, f_cat, f_prio)
            f_user = c1.selectbox("👤 Responsável", users, key=key_filter_user)

            cats = ["Todas"] + sorted(df_demandas['categoria'].dropna().unique().tolist())
            f_cat = c2.selectbox("🏷️ Categoria", cats, key="planner_filter_cat")

            priorities = ["Todas", "Alta", "Média", "Baixa"]
            f_prio = c3.selectbox("🚨 Prioridade", priorities, key="planner_filter_prio")

        st.divider()

        # ---------------------------------------------------------
        # B. APLICAÇÃO DOS FILTROS (Agora as variáveis já existem)
        # ---------------------------------------------------------
        df_view = df_demandas.copy()

        # Garante coluna prioridade
        if 'prioridade' not in df_view.columns: df_view['prioridade'] = 'Média'

        # Aplica filtro de usuário
        if f_user != "Todos":
            df_view = df_view[df_view['assigned_to'].astype(str) == f_user]

        # Aplica filtro de categoria
        if f_cat != "Todas":
            df_view = df_view[df_view['categoria'] == f_cat]

        # Aplica filtro de prioridade
        if f_prio != "Todas":
            df_view = df_view[df_view['prioridade'] == f_prio]

        # ---------------------------------------------------------
        # C. DEFINIÇÃO VISUAL
        # ---------------------------------------------------------
        cat_colors = {
            "Facilities/Eletromecânica": "#E1BEE7",
            "Manutenção de rede": "#BBDEFB",
            "Tratamento": "#C8E6C9",
            "Tratamento (Laboratório)": "#FFECB3"
        }

        kanban_columns = {
            "A Fazer": ["Aberta", "Pendente"],
            "Em Andamento": ["Em Atendimento", "Em Processamento", "Aguardando Aprovação"],
            "Concluído": ["Finalizado", "Entregue", "Concluída", "Fechada"]
        }

        # CSS para etiquetas
        st.markdown("""
        <style>
        .planner-tag {
            padding: 2px 8px; border-radius: 12px; font-size: 0.75rem; font-weight: bold; color: #444; display: inline-block; margin-bottom: 5px;
        }
        .planner-date {
            font-size: 0.8rem; color: #666; display: flex; align-items: center; gap: 4px;
        }
        </style>
        """, unsafe_allow_html=True)

        cols = st.columns(len(kanban_columns))

        # ---------------------------------------------------------
        # D. RENDERIZAÇÃO DAS COLUNAS
        # ---------------------------------------------------------
        for i, (col_name, status_list) in enumerate(kanban_columns.items()):
            with cols[i]:
                tasks = df_view[df_view['status_demanda'].isin(status_list)]
                st.markdown(f"#### {col_name} <span style='font-size:0.8em; color:gray'>({len(tasks)})</span>",
                            unsafe_allow_html=True)

                for _, row in tasks.iterrows():
                    bg_color = cat_colors.get(row.get('categoria'), "#F0F0F0")
                    prio = row.get('prioridade', 'Média')
                    prio_icon = "🔴" if prio == "Alta" else ("🟡" if prio == "Média" else "🟢")

                    with st.container(border=True):
                        # 1. Etiquetas
                        st.markdown(f"""
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <span class="planner-tag" style="background-color:{bg_color}">{row.get('categoria', 'Geral')}</span>
                            <span title="Prioridade {prio}">{prio_icon}</span>
                        </div>
                        """, unsafe_allow_html=True)

                        # 2. Descrição
                        desc = row.get('descricao_necessidade', '')
                        short_desc = (desc[:75] + '...') if len(desc) > 75 else desc
                        st.markdown(f"**{short_desc}**")

                        # 3. Rodapé (Data e Usuário) - COM PROTEÇÃO DE ERRO
                        data_fmt = row['created_at'].strftime('%d/%b') if pd.notnull(row.get('created_at')) else ""

                        # Lógica segura para pegar o nome
                        raw_assigned = row.get('assigned_to')
                        if isinstance(raw_assigned, str) and raw_assigned.strip():
                            atribuido = raw_assigned.split(' ')[0]
                        else:
                            atribuido = "Ninguém"

                        st.markdown(f"""
                        <div style="display:flex; justify-content:space-between; align-items:center; margin-top:8px; margin-bottom:8px;">
                            <div class="planner-date">📅 {data_fmt}</div>
                            <div class="planner-date">👤 {atribuido}</div>
                        </div>
                        """, unsafe_allow_html=True)

                        # 4. Ações
                        c_move, c_edit = st.columns([2, 1])

                        current_status = row.get('status_demanda', 'Aberta')
                        all_statuses = ["Aberta", "Em Atendimento", "Concluída", "Cancelado"]

                        target_status = c_move.selectbox(
                            "Mover",
                            all_statuses,
                            index=all_statuses.index(current_status) if current_status in all_statuses else 0,
                            key=f"move_{row['id']}",
                            label_visibility="collapsed"
                        )

                        if target_status != current_status:
                            self.db.update_doc("demandas", row['id'],
                                               {"status_demanda": target_status},
                                               st.session_state.username)
                            st.toast(f"Movido para {target_status}")
                            time.sleep(0.5)
                            st.rerun()

                        if c_edit.button("✏️", key=f"edit_btn_{row['id']}", help="Editar detalhes"):
                            self._set_edit_state("demandas", row.to_dict())
                            st.rerun()

    def render_logs_tab(self):
        st.header("🛡️ Registros de Atividades do Sistema")

        # OTIMIZAÇÃO CRÍTICA: Carrega logs do cache
        logs_df = get_cached_docs(self.db, "audit_logs")

        if logs_df.empty:
            st.info("Nenhum registro de atividade encontrado.")
            return

        col1, col2 = st.columns(2)
        with col1:
            # Verifica se a coluna existe antes de usar e converte para string para evitar erros de tipo
            if 'username' in logs_df.columns:
                unique_users = sorted(logs_df['username'].astype(str).unique().tolist())
            else:
                unique_users = []
            users = ["Todos"] + unique_users
            selected_user = st.selectbox("Filtrar por Usuário", users)

        with col2:
            if 'action' in logs_df.columns:
                unique_actions = sorted(logs_df['action'].astype(str).unique().tolist())
            else:
                unique_actions = []
            actions = ["Todos"] + unique_actions
            selected_action = st.selectbox("Filtrar por Ação", actions)

        filtered_df = logs_df.copy()
        if selected_user != "Todos":
            filtered_df = filtered_df[filtered_df['username'] == selected_user]
        if selected_action != "Todos":
            filtered_df = filtered_df[filtered_df['action'] == selected_action]

        # Seleciona apenas colunas existentes para evitar KeyError
        cols_to_show = ['timestamp', 'username', 'action', 'details']
        available_cols = [c for c in cols_to_show if c in filtered_df.columns]

        df_para_exibir = filtered_df[available_cols].copy()

        if 'details' in df_para_exibir.columns:
            df_para_exibir['details'] = df_para_exibir['details'].astype(str)

        st.dataframe(
            df_para_exibir,
            use_container_width=True,
            column_config={
                "timestamp": st.column_config.DatetimeColumn("Data e Hora", format="DD/MM/YYYY - HH:mm:ss"),
                "username": "Usuário",
                "action": "Ação",
                "details": "Detalhes"
            },
            hide_index=True
        )

    @st.dialog("🔔 Notificações")
    def render_notifications_modal(self):
        if not st.session_state.notifications_list:
            st.info("Nenhuma notificação encontrada.")
            if st.button("Fechar"):
                st.session_state.show_notifications = False
                st.rerun()
            return

        if st.button("Marcar todas como lidas"):
            batch = self.db.db.batch()  # Usa batch para economizar writes
            ids_to_ignore = []

            for notif in st.session_state.notifications_list:
                # Se for notificação normal (tem ID do banco)
                if 'id' in notif and not notif['id'].startswith('admin_approval'):
                    doc_ref = self.db.db.collection("notifications").document(notif['id'])
                    batch.update(doc_ref, {"read": True})
                    ids_to_ignore.append(notif['id'])

            batch.commit()

            # Atualiza LOCALMENTE sem limpar cache global
            st.session_state.ignored_notifications.update(ids_to_ignore)
            st.session_state.show_notifications = False
            st.rerun()

        st.divider()

        for i, notif in enumerate(st.session_state.notifications_list):
            with st.container(border=True):
                col_txt, col_btn = st.columns([0.85, 0.15])

                with col_txt:
                    icone = "👤" if notif.get('type') == 'admin_approval' else "💬"
                    st.markdown(f"**{icone} {notif.get('author', 'Sistema')}**")
                    st.write(notif.get('message', ''))
                    if 'timestamp' in notif:
                        ts = notif['timestamp']
                        if hasattr(ts, 'strftime'):
                            st.caption(f"Em: {ts.strftime('%d/%m/%Y %H:%M')}")

                with col_btn:
                    # AÇÃO: APROVAR USUÁRIO
                    if notif.get('type') == 'admin_approval':
                        user_id = notif['id'].replace("admin_approval_", "")
                        if st.button("✅", key=f"ok_{i}"):
                            # Aqui precisamos usar o update normal pois altera tabela de usuários (importante refletir globalmente)
                            # Mas para não travar, podemos fazer update direto e ignorar notificação
                            self.db.db.collection("users").document(user_id).update({"status": "active"})
                            self.db.log_action("User Approved", st.session_state.username, {"target": user_id})

                            # Remove da vista imediatamente
                            st.session_state.ignored_notifications.add(notif['id'])
                            st.rerun()

                    # AÇÃO: MARCAR COMO LIDA (Notificação normal)
                    elif 'id' in notif:
                        if st.button("✖️", key=f"del_{i}"):
                            # 1. Update silencioso no Firebase (sem clear_cache)
                            self.db.db.collection("notifications").document(notif['id']).update({"read": True})

                            # 2. Update na memória local
                            st.session_state.ignored_notifications.add(notif['id'])

                            # 3. Rerun para atualizar a UI instantaneamente
                            st.rerun()

# -----------------------------------------------------------------------------
# 4. PONTO DE ENTRADA DA APLICAÇÃO
# -----------------------------------------------------------------------------

if __name__ == "__main__":
    try:
        if "firebase_credentials" not in st.secrets:
            st.error("Credenciais do Firebase não encontradas! Verifique seu arquivo secrets.toml.");
            st.stop()
        db_service = get_db_service(dict(st.secrets["firebase_credentials"]))
        auth_service = AuthService(db_service)
        app = ViewManager(auth_service, db_service)
        app.run()
    except Exception as e:
        st.error("Ocorreu um erro crítico na aplicação.")
        st.exception(e)
        logger.critical(f"Erro crítico na aplicação: {e}", exc_info=True)